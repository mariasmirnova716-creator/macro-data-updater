from __future__ import annotations

import io
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = PROJECT_ROOT / "data" / "sheet_01_gdp_data"

ANNUAL_OUTPUT = (
    DATA_DIR / "russia_labor_productivity_annual.csv"
)

OPERATIONAL_OUTPUT = (
    DATA_DIR / "russia_labor_productivity_operational.csv"
)

COMBINED_OUTPUT = (
    DATA_DIR / "russia_labor_productivity.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR / "russia_labor_productivity_source.txt"
)

ROSSTAT_PAGE = "https://rosstat.gov.ru/statistics/accounts"

TARGET_TEXT = "в целом по экономике российской федерации"


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("«", '"')
    text = text.replace("»", '"')

    text = text.strip().lower()

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text


def as_number(value):

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    text = text.replace(",", ".")

    text = re.sub(
        r"[^\d.\-]",
        "",
        text,
    )

    if text in {
        "",
        "-",
        ".",
        "-.",
    }:
        return None

    try:
        return float(text)

    except ValueError:
        return None


def detect_year(value) -> int | None:

    text = clean_text(value)

    match = re.search(
        r"\b(20\d{2})\b",
        text,
    )

    if match:
        return int(
            match.group(1)
        )

    return None

def request_bytes(url: str) -> bytes:
    """
    Robust Rosstat downloader.

    1. Tries normal verified HTTPS.
    2. Retries on temporary connection/server errors.
    3. Tries rosstat.gov.ru and www.rosstat.gov.ru.
    4. Falls back to verify=False only if SSL verification fails.
    """

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 Chrome/126 Safari/537.36"
        )
    }

    urls_to_try = [url]

    if "://rosstat.gov.ru/" in url:
        urls_to_try.append(
            url.replace(
                "://rosstat.gov.ru/",
                "://www.rosstat.gov.ru/",
            )
        )

    elif "://www.rosstat.gov.ru/" in url:
        urls_to_try.append(
            url.replace(
                "://www.rosstat.gov.ru/",
                "://rosstat.gov.ru/",
            )
        )

    last_error = None

    for current_url in urls_to_try:

        for attempt in range(1, 4):

            print(
                f"Request attempt {attempt}/3: "
                f"{current_url}"
            )

            try:
                response = requests.get(
                    current_url,
                    timeout=90,
                    headers=headers,
                    verify=certifi.where(),
                )

                response.raise_for_status()

                return response.content

            except requests.exceptions.SSLError as exc:

                print(
                    "WARNING: SSL verification failed."
                )
                print(
                    "Retrying without certificate verification."
                )

                warnings.simplefilter(
                    "ignore",
                    InsecureRequestWarning,
                )

                try:
                    response = requests.get(
                        current_url,
                        timeout=90,
                        headers=headers,
                        verify=False,
                    )

                    response.raise_for_status()

                    return response.content

                except requests.exceptions.RequestException as inner_exc:
                    last_error = inner_exc

            except requests.exceptions.RequestException as exc:

                last_error = exc

                print(
                    f"WARNING: request failed: {exc}"
                )

    raise RuntimeError(
        "Rosstat request failed after all retries.\n"
        f"Original URL: {url}\n"
        f"Last error: {last_error}"
    )


def download_xlsx(url: str) -> bytes:

    content = request_bytes(url)

    if not content.startswith(b"PK"):

        raise RuntimeError(
            "Downloaded Rosstat file "
            "does not look like XLSX:\n"
            + url
        )

    return content


def url_filename(url: str) -> str:

    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


# ============================================================
# FIND CURRENT ROSSTAT FILES
# ============================================================

def find_current_rosstat_files() -> dict[str, str]:
    """
    Find only XLSX links from the Rosstat section
    'Индекс производительности труда'.

    We locate the section title in raw HTML and inspect
    only the nearby fragment until the next major section.
    """

    print("Opening Rosstat page:")
    print(ROSSTAT_PAGE)

    html = request_bytes(
        ROSSTAT_PAGE
    ).decode(
        "utf-8",
        errors="ignore",
    )

    html_lower = html.lower()

    # --------------------------------------------------------
    # 1. Find the labor productivity section
    # --------------------------------------------------------

    target_phrase = "индекс производительности труда"

    start = html_lower.find(
        target_phrase
    )

    if start < 0:
        raise RuntimeError(
            "Section 'Индекс производительности труда' "
            "was not found in Rosstat HTML."
        )

    print(
        "Labor productivity section found."
    )

    # --------------------------------------------------------
    # 2. Take only a limited fragment after the section title
    #
    # 20k chars is enough for the accordion content,
    # but avoids scanning the whole National Accounts page.
    # --------------------------------------------------------

    fragment = html[
        start:
        min(
            len(html),
            start + 20000,
        )
    ]

    fragment_soup = BeautifulSoup(
        fragment,
        "html.parser",
    )

    candidates = []

    for tag in fragment_soup.find_all(
        "a",
        href=True,
    ):

        href = urljoin(
            ROSSTAT_PAGE,
            tag["href"],
        )

        if not url_filename(
            href
        ).endswith(".xlsx"):
            continue

        candidates.append(
            href
        )

    candidates = list(
        dict.fromkeys(
            candidates
        )
    )

    print(
        "\nLabor productivity XLSX candidates:"
    )

    for url in candidates:
        print(url)

    if not candidates:
        raise RuntimeError(
            "No XLSX links found near "
            "the labor productivity section."
        )

    # --------------------------------------------------------
    # 3. Inspect only these few files by content
    # --------------------------------------------------------

    annual_url = None
    operational_url = None

    for url in candidates:

        print(
            f"\nChecking: {url_filename(url)}"
        )

        content = download_xlsx(
            url
        )

        detected = identify_workbook_type(
            content
        )

        print(
            f"Detected type: {detected}"
        )

        if detected == "annual":
            annual_url = url

        elif detected == "operational":
            operational_url = url

        if (
            annual_url is not None
            and operational_url is not None
        ):
            break

    if annual_url is None:
        raise RuntimeError(
            "Annual labor productivity XLSX "
            "was not found."
        )

    if operational_url is None:
        raise RuntimeError(
            "Operational labor productivity XLSX "
            "was not found."
        )

    print(
        "\nCurrent annual productivity:"
    )
    print(
        annual_url
    )

    print(
        "\nCurrent operational productivity:"
    )
    print(
        operational_url
    )

    return {
        "annual": annual_url,
        "operational": operational_url,
    }


# ============================================================
# WORKBOOK IDENTIFICATION
# ============================================================

def workbook_text(
    content: bytes,
    max_rows: int = 15,
    max_cols: int = 20,
) -> str:
    """
    Read visible text from the top part of all sheets.
    Used only to identify workbook type.
    """

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=True,
    )

    pieces = []

    for ws in wb.worksheets:

        pieces.append(
            clean_text(ws.title)
        )

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(
                ws.max_row,
                max_rows,
            ),
            min_col=1,
            max_col=min(
                ws.max_column,
                max_cols,
            ),
            values_only=True,
        ):

            for value in row:

                text = clean_text(value)

                if text:
                    pieces.append(text)

    return " ".join(
        pieces
    )


def identify_workbook_type(
    content: bytes,
) -> str | None:
    """
    annual:
      title contains labor productivity,
      typically years 2018, 2019, 2020...
      and wording "% к предыдущему году"

    operational:
      title explicitly contains "оперативный",
      periods such as:
        1 квартал
        1 полугодие
        9 месяцев
        год
    """

    text = workbook_text(
        content
    )

    if (
        "производительности труда"
        not in text
    ):
        return None

    if (
        "оператив" in text
        and "1 полугодие" in text
        and "9 месяцев" in text
    ):
        return "operational"

    if (
        "к предыдущему году" in text
        or (
            "2018" in text
            and "2019" in text
            and "2020" in text
        )
    ):
        return "annual"

    return None


def choose_workbook_by_content(
    urls: list[str],
    workbook_type: str,
) -> str | None:

    for url in urls:

        try:

            content = download_xlsx(
                url
            )

            detected = identify_workbook_type(
                content
            )

            print(
                f"Checked {url_filename(url)} "
                f"-> {detected}"
            )

            if detected == workbook_type:
                return url

        except Exception as exc:

            print(
                "Skipping candidate:",
                url,
            )

            print(
                "Reason:",
                exc,
            )

    return None


def search_all_xlsx_by_content(
    urls: list[str],
    workbook_type: str,
) -> str | None:
    """
    Final fallback.

    To avoid downloading hundreds of irrelevant files,
    inspect only URLs whose filename or surrounding page
    position might be relevant first.

    If Rosstat changes the publication link structure
    completely, content itself remains authoritative.
    """

    checked = set()

    for url in urls:

        if url in checked:
            continue

        checked.add(url)

        name = url_filename(url)

        # Skip obviously unrelated large families
        # where possible.
        obvious_unrelated = [
            "vvp_",
            "vds_",
            "vrp_",
            "schet",
            "tabl",
        ]

        if any(
            item in name
            for item in obvious_unrelated
        ):
            continue

        try:

            content = download_xlsx(
                url
            )

            detected = identify_workbook_type(
                content
            )

            if detected == workbook_type:

                print(
                    "\nFound by workbook content:"
                )

                print(url)

                return url

        except Exception:
            continue

    return None


# ============================================================
# FIND TARGET ROW
# ============================================================

def row_contains_target(
    ws,
    row_number: int,
) -> bool:

    for col in range(
        1,
        min(
            ws.max_column,
            10,
        ) + 1,
    ):

        text = clean_text(
            ws.cell(
                row_number,
                col,
            ).value
        )

        if (
            TARGET_TEXT in text
        ):
            return True

    return False


def find_target_row(
    ws,
) -> int | None:

    for row in range(
        1,
        ws.max_row + 1,
    ):

        if row_contains_target(
            ws,
            row,
        ):
            return row

    return None


def find_sheet_with_target(
    wb,
):

    for ws in wb.worksheets:

        row = find_target_row(
            ws
        )

        if row is not None:
            return ws, row

    raise RuntimeError(
        "Row 'В целом по экономике "
        "Российской Федерации' "
        "was not found in workbook."
    )


# ============================================================
# ANNUAL PARSER
# ============================================================

def parse_annual(
    content: bytes,
) -> pd.DataFrame:

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=True,
    )

    ws, target_row = find_sheet_with_target(
        wb
    )

    print(
        "\nAnnual workbook:"
    )

    print(
        f"Sheet: {ws.title}"
    )

    print(
        f"Target row: {target_row}"
    )

    # --------------------------------------------------------
    # Find header row containing the largest number of years.
    # --------------------------------------------------------

    best_year_row = None
    best_year_count = 0

    search_start = max(
        1,
        target_row - 12,
    )

    for row in range(
        search_start,
        target_row,
    ):

        count = 0

        for col in range(
            1,
            ws.max_column + 1,
        ):

            year = detect_year(
                ws.cell(
                    row,
                    col,
                ).value
            )

            if year is not None:
                count += 1

        if count > best_year_count:

            best_year_count = count
            best_year_row = row

    if (
        best_year_row is None
        or best_year_count < 2
    ):

        raise RuntimeError(
            "Could not detect year header "
            "in annual productivity workbook."
        )

    print(
        f"Year row: {best_year_row} "
        f"({best_year_count} years)"
    )

    rows = []

    for col in range(
        1,
        ws.max_column + 1,
    ):

        year = detect_year(
            ws.cell(
                best_year_row,
                col,
            ).value
        )

        if year is None:
            continue

        value = as_number(
            ws.cell(
                target_row,
                col,
            ).value
        )

        if value is None:
            continue

        rows.append(
            {
                "year": year,
                "labor_productivity_index": value,
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        raise RuntimeError(
            "No annual labor productivity "
            "values extracted."
        )

    df = (
        df
        .sort_values("year")
        .drop_duplicates(
            "year",
            keep="last",
        )
        .reset_index(drop=True)
    )

    df[
        "labor_productivity_index"
    ] = pd.to_numeric(
        df[
            "labor_productivity_index"
        ],
        errors="coerce",
    ).round(1)

    return df


# ============================================================
# OPERATIONAL PARSER
# ============================================================

def detect_period(
    value,
) -> str | None:

    text = clean_text(value)

    # Remove footnote symbols / digits.
    text = re.sub(
        r"[¹²³⁴⁵⁶⁷⁸⁹⁰]+",
        "",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    mappings = {
        "1 квартал": "q1",
        "i квартал": "q1",
        "1 полугодие": "h1",
        "i полугодие": "h1",
        "9 месяцев": "m9",
        "год": "year",
    }

    return mappings.get(
        text
    )


def parse_operational(
    content: bytes,
) -> pd.DataFrame:

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=True,
    )

    ws, target_row = find_sheet_with_target(
        wb
    )

    print(
        "\nOperational workbook:"
    )

    print(
        f"Sheet: {ws.title}"
    )

    print(
        f"Target row: {target_row}"
    )

    # --------------------------------------------------------
    # Detect period row.
    # Should contain:
    # 1 квартал / 1 полугодие / 9 месяцев / год
    # --------------------------------------------------------

    best_period_row = None
    best_period_count = 0

    search_start = max(
        1,
        target_row - 12,
    )

    for row in range(
        search_start,
        target_row,
    ):

        count = 0

        for col in range(
            1,
            ws.max_column + 1,
        ):

            period = detect_period(
                ws.cell(
                    row,
                    col,
                ).value
            )

            if period is not None:
                count += 1

        if count > best_period_count:

            best_period_count = count
            best_period_row = row

    if (
        best_period_row is None
        or best_period_count < 2
    ):

        raise RuntimeError(
            "Could not detect period header "
            "in operational workbook."
        )

    print(
        f"Period row: {best_period_row} "
        f"({best_period_count} periods)"
    )

    # --------------------------------------------------------
    # Find year header above periods.
    #
    # Years are merged across four columns.
    # Therefore year appears only in the first cell
    # of each 4-column block.
    # --------------------------------------------------------

    year_row = None
    best_year_count = 0

    for row in range(
        max(
            1,
            best_period_row - 5,
        ),
        best_period_row,
    ):

        count = 0

        for col in range(
            1,
            ws.max_column + 1,
        ):

            year = detect_year(
                ws.cell(
                    row,
                    col,
                ).value
            )

            if year is not None:
                count += 1

        if count > best_year_count:

            best_year_count = count
            year_row = row

    if year_row is None:

        raise RuntimeError(
            "Could not detect year header "
            "in operational workbook."
        )

    print(
        f"Year row: {year_row}"
    )

    # --------------------------------------------------------
    # Parse left -> right.
    #
    # Because year cells are merged across four columns,
    # carry the latest detected year forward.
    # --------------------------------------------------------

    rows = []

    current_year = None

    for col in range(
        1,
        ws.max_column + 1,
    ):

        detected_year = detect_year(
            ws.cell(
                year_row,
                col,
            ).value
        )

        if detected_year is not None:

            current_year = detected_year

        period = detect_period(
            ws.cell(
                best_period_row,
                col,
            ).value
        )

        if period is None:
            continue

        if current_year is None:
            continue

        value = as_number(
            ws.cell(
                target_row,
                col,
            ).value
        )

        # Future periods are blank:
        # e.g. 2026 H1 / 9M / year.
        # Do not create rows for blanks.
        if value is None:
            continue

        rows.append(
            {
                "year": current_year,
                "period": period,
                "labor_productivity_operational": value,
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        raise RuntimeError(
            "No operational labor productivity "
            "values extracted."
        )

    period_order = {
        "q1": 1,
        "h1": 2,
        "m9": 3,
        "year": 4,
    }

    df["_period_order"] = (
        df["period"]
        .map(period_order)
    )

    df = (
        df
        .sort_values(
            [
                "year",
                "_period_order",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "period",
            ],
            keep="last",
        )
        .drop(
            columns=[
                "_period_order",
            ]
        )
        .reset_index(drop=True)
    )

    df[
        "labor_productivity_operational"
    ] = pd.to_numeric(
        df[
            "labor_productivity_operational"
        ],
        errors="coerce",
    ).round(1)

    return df


# ============================================================
# VALIDATION
# ============================================================

def validate_annual(
    df: pd.DataFrame,
):

    if df.empty:

        raise RuntimeError(
            "Annual output is empty."
        )

    if df[
        "year"
    ].duplicated().any():

        raise RuntimeError(
            "Duplicate years "
            "in annual output."
        )

    bad = df[
        ~df[
            "labor_productivity_index"
        ].between(
            50,
            150,
        )
    ]

    if not bad.empty:

        raise RuntimeError(
            "Suspicious annual labor "
            "productivity values:\n"
            + bad.to_string(
                index=False
            )
        )

    # Our current official history
    # visibly contains 2018 onward.
    if (
        int(
            df["year"].min()
        )
        > 2018
    ):

        raise RuntimeError(
            "Annual productivity history "
            "starts later than expected."
        )


def validate_operational(
    df: pd.DataFrame,
):

    if df.empty:

        raise RuntimeError(
            "Operational output is empty."
        )

    if df.duplicated(
        [
            "year",
            "period",
        ]
    ).any():

        raise RuntimeError(
            "Duplicate year-period "
            "in operational output."
        )

    allowed_periods = {
        "q1",
        "h1",
        "m9",
        "year",
    }

    invalid_periods = set(
        df["period"]
    ) - allowed_periods

    if invalid_periods:

        raise RuntimeError(
            "Unexpected operational periods: "
            + str(
                invalid_periods
            )
        )

    bad = df[
        ~df[
            "labor_productivity_operational"
        ].between(
            50,
            150,
        )
    ]

    if not bad.empty:

        raise RuntimeError(
            "Suspicious operational labor "
            "productivity values:\n"
            + bad.to_string(
                index=False
            )
        )

    # The new operational publication
    # currently starts in 2025.
    if (
        int(
            df["year"].min()
        )
        > 2025
    ):

        raise RuntimeError(
            "Operational productivity "
            "history unexpectedly starts "
            "after 2025."
        )


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 72)
    print(
        "Russia labor productivity updater "
        "— Sheet 1"
    )
    print("=" * 72)

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    urls = find_current_rosstat_files()

    print(
        "\nDownloading current Rosstat workbooks..."
    )

    annual_content = download_xlsx(
        urls["annual"]
    )

    operational_content = download_xlsx(
        urls["operational"]
    )

    print(
        "\nParsing annual productivity..."
    )

    annual = parse_annual(
        annual_content
    )

    print(
        "\nParsing operational productivity..."
    )

    operational = parse_operational(
        operational_content
    )

    print(
        "\nValidating..."
    )

    validate_annual(
        annual
    )

    validate_operational(
        operational
    )

    annual_columns = [
        "year",
        "labor_productivity_index",
    ]

    operational_columns = [
        "year",
        "period",
        "labor_productivity_operational",
    ]

    annual[
        annual_columns
    ].to_csv(
        ANNUAL_OUTPUT,
        index=False,
        encoding="utf-8",
    )

    operational[
        operational_columns
    ].to_csv(
        OPERATIONAL_OUTPUT,
        index=False,
        encoding="utf-8",
    )
    # ========================================================
    # COMBINED CSV FOR EXCEL / POWER QUERY
    # ========================================================

    annual_combined = annual.rename(
        columns={
            "labor_productivity_index": "value"
        }
    ).copy()

    annual_combined["series"] = "annual"
    annual_combined["period"] = "year"

    annual_combined = annual_combined[
        ["series", "year", "period", "value"]
    ]

    operational_combined = operational.rename(
        columns={
            "labor_productivity_operational": "value"
        }
    ).copy()

    operational_combined["series"] = "operational"

    operational_combined = operational_combined[
        ["series", "year", "period", "value"]
    ]

    combined = pd.concat(
        [
            annual_combined,
            operational_combined,
        ],
        ignore_index=True,
    )

    combined.to_csv(
        COMBINED_OUTPUT,
        index=False,
        encoding="utf-8",
    )

    retrieved_at = datetime.now(
        timezone.utc
    ).isoformat()

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                (
                    "Russia labor productivity "
                    "— Sheet 1"
                ),
                f"updated_utc={retrieved_at}",
                f"rosstat_page={ROSSTAT_PAGE}",
                "",
                (
                    "annual="
                    + urls["annual"]
                ),
                (
                    "operational="
                    + urls["operational"]
                ),
                "",
                (
                    "annual_definition="
                    "% to previous year"
                ),
                (
                    "operational_definition="
                    "% to corresponding period "
                    "of previous year"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print("DONE")

    print("=" * 72)

    print(
        f"\nAnnual output:\n"
        f"{ANNUAL_OUTPUT}"
    )

    print(
        f"\nOperational output:\n"
        f"{OPERATIONAL_OUTPUT}"
    )

    print(
        f"\nSource info:\n"
        f"{SOURCE_INFO_FILE}"
    )

    print(
        "\nAnnual data:"
    )

    print(
        annual[
            annual_columns
        ].to_string(
            index=False
        )
    )

    print(
        "\nOperational data:"
    )

    print(
        operational[
            operational_columns
        ].to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()