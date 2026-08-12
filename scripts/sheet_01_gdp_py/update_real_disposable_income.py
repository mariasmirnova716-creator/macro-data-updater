from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = PROJECT_ROOT / "data" / "sheet_01_gdp_data"

OUTPUT_FILE = (
    DATA_DIR
    / "russia_real_disposable_income.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_real_disposable_income_source.txt"
)


# ============================================================
# ROSSTAT
# ============================================================

ROSSTAT_PAGE = "https://rosstat.gov.ru/folder/13397"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def as_number(value):

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):

        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = text.replace(",", ".")

    text = re.sub(
        r"[^\d.\-]",
        "",
        text,
    )

    if text in {
        "",
        "-",
        ".",
        "-.",
    }:
        return None

    try:
        return float(text)

    except ValueError:
        return None


def url_filename(url: str) -> str:

    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:

        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:

        print(
            "WARNING: SSL verification failed."
        )

        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND CURRENT ROSSTAT FILE
# ============================================================

def find_current_rosstat_file() -> str:

    print(
        "\nOpening Rosstat living standards page:"
    )

    print(
        ROSSTAT_PAGE
    )

    with requests.Session() as session:

        response = request(
            session,
            "GET",
            ROSSTAT_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    candidates = []

    for tag in soup.find_all(
        "a",
        href=True,
    ):

        href = urljoin(
            ROSSTAT_PAGE,
            tag["href"],
        )

        filename = url_filename(
            href
        )

        if not filename.endswith(
            (".xlsx", ".xls")
        ):
            continue

        text_parts = [
            tag.get_text(
                " ",
                strip=True,
            )
        ]

        parent = tag.parent

        for _ in range(4):

            if parent is None:
                break

            text_parts.append(
                parent.get_text(
                    " ",
                    strip=True,
                )
            )

            parent = parent.parent

        surrounding_text = clean_text(
            " ".join(text_parts)
        )

        # Main filter by Rosstat indicator title.
        if (
            "реальные" in surrounding_text
            and
            "располагаемые" in surrounding_text
            and
            "денежные доходы" in surrounding_text
            and
            "российской федерации" in surrounding_text
        ):

            candidates.append(
                href
            )

        # Known federal filename family fallback.
        elif (
            "urov" in filename
            and
            (
                "12kv" in filename
                or "rrdd" in filename
            )
        ):

            candidates.append(
                href
            )

    candidates = list(
        dict.fromkeys(
            candidates
        )
    )

    print(
        "Income workbook candidates:",
        len(candidates),
    )

    if not candidates:

        raise RuntimeError(
            "Could not find Rosstat workbook "
            "for real disposable income."
        )

    # Prefer candidate that can actually be parsed.
    for url in candidates:

        print(
            "Checking:",
            url,
        )

        try:

            content = download_excel(
                url
            )

            wb = load_workbook(
                io.BytesIO(content),
                data_only=True,
                read_only=False,
            )

            sheet_names = [
                clean_text(name)
                for name in wb.sheetnames
            ]

            if any(
                "ррдд" in name
                for name in sheet_names
            ):
                print(
                    "Selected Rosstat workbook:"
                )

                print(
                    url
                )

                return url

        except Exception:
            continue

    raise RuntimeError(
        "Rosstat income workbook candidates "
        "were found, but none could be confirmed."
    )


def download_excel(
    url: str,
) -> bytes:

    with requests.Session() as session:

        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    # XLSX is a ZIP container.
    if content.startswith(b"PK"):
        return content

    raise RuntimeError(
        "Downloaded Rosstat file "
        "does not look like XLSX:\n"
        + url
    )


# ============================================================
# PARSER
# ============================================================

def detect_period(
    value,
) -> str | None:

    text = clean_text(
        value
    )

    if "1 кварт" in text:
        return "q1"

    if "2 кварт" in text:
        return "q2"

    if "3 кварт" in text:
        return "q3"

    if "4 кварт" in text:
        return "q4"

    if text == "год":
        return "year"

    return None


def detect_year(
    value,
) -> int | None:

    text = clean_text(
        value
    )

    match = re.search(
        r"\b(20\d{2})\b",
        text,
    )

    if match:
        return int(
            match.group(1)
        )

    return None


def parse_real_disposable_income(
    content: bytes,
) -> pd.DataFrame:

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=False,
    )

    # Prefer explicit Rosstat sheet name.
    target_ws = None

    for ws in wb.worksheets:

        sheet_name = clean_text(
            ws.title
        )

        if (
            "ррдд" in sheet_name
            and
            "рдд" in sheet_name
        ):

            target_ws = ws
            break

    if target_ws is None:

        for ws in wb.worksheets:

            full_text = []

            for row in range(
                1,
                min(ws.max_row, 20) + 1,
            ):

                for col in range(
                    1,
                    min(ws.max_column, 10) + 1,
                ):

                    value = ws.cell(
                        row,
                        col,
                    ).value

                    if value is not None:
                        full_text.append(
                            clean_text(value)
                        )

            text = " ".join(
                full_text
            )

            if (
                "реальные располагаемые" in text
                and
                "соответствующему периоду" in text
            ):

                target_ws = ws
                break

    if target_ws is None:

        raise RuntimeError(
            "Could not identify Rosstat "
            "real disposable income sheet."
        )

    ws = target_ws

    print(
        "\nParsing sheet:",
        ws.title,
    )

    # ========================================================
    # Find YoY column for real disposable income
    # ========================================================

    target_col = None

    # Usually header is within first ~10 rows.
    for col in range(
        1,
        ws.max_column + 1,
    ):

        parts = []

        for row in range(
            1,
            min(ws.max_row, 15) + 1,
        ):

            value = ws.cell(
                row,
                col,
            ).value

            if value is not None:

                parts.append(
                    clean_text(value)
                )

        header = " ".join(
            parts
        )

        if (
            "соответствующему" in header
            and
            "периоду" in header
        ):

            # We want the disposable-income section,
            # which is the right-hand block.
            target_col = col

    if target_col is None:

        raise RuntimeError(
            "Could not find real disposable "
            "income YoY column."
        )

    print(
        "Disposable-income YoY column:",
        target_col,
    )

    # ========================================================
    # Parse years + periods
    # ========================================================

    rows = []

    current_year = None

    for row in range(
        1,
        ws.max_row + 1,
    ):

        # Search the whole row for year header.
        detected_year = None

        for col in range(
            1,
            ws.max_column + 1,
        ):

            year = detect_year(
                ws.cell(
                    row,
                    col,
                ).value
            )

            if year is not None:
                detected_year = year
                break

        # A year header must not itself be a quarter row.
        period = detect_period(
            ws.cell(
                row,
                1,
            ).value
        )

        if (
            detected_year is not None
            and
            period is None
        ):

            current_year = detected_year

        if current_year is None:
            continue

        period = detect_period(
            ws.cell(
                row,
                1,
            ).value
        )

        if period is None:
            continue

        value = as_number(
            ws.cell(
                row,
                target_col,
            ).value
        )

        if value is None:
            continue

        rows.append(
            {
                "year": current_year,
                "period": period,
                "real_disposable_income_yoy": value,
                "source": "rosstat_current",
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        raise RuntimeError(
            "No real disposable income "
            "observations were extracted."
        )

    df = (
        df
        .sort_values(
            [
                "year",
                "period",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "period",
            ],
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    df[
        "real_disposable_income_yoy"
    ] = pd.to_numeric(
        df[
            "real_disposable_income_yoy"
        ],
        errors="coerce",
    ).round(1)

    return df


# ============================================================
# PRESERVE HISTORY
# ============================================================

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():

        return current

    print(
        "\nLoading existing income history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required = {
        "year",
        "period",
        "real_disposable_income_yoy",
    }

    if not required.issubset(
        existing.columns
    ):

        raise RuntimeError(
            "Existing real disposable income CSV "
            "has unexpected columns."
        )

    if "source" not in existing.columns:

        existing[
            "source"
        ] = "existing_history"

    existing_keys = set(
        zip(
            existing["year"].astype(int),
            existing["period"].astype(str),
        )
    )

    current_new = current[
        ~current.apply(
            lambda row:
            (
                int(row["year"]),
                str(row["period"]),
            )
            in existing_keys,
            axis=1,
        )
    ].copy()

    result = pd.concat(
        [
            existing,
            current_new,
        ],
        ignore_index=True,
        sort=False,
    )

    return (
        result
        .sort_values(
            [
                "year",
                "period",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "period",
            ],
            keep="first",
        )
        .reset_index(
            drop=True
        )
    )


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:
        raise RuntimeError(
            "Disposable income output is empty."
        )

    if df.duplicated(
        [
            "year",
            "period",
        ]
    ).any():

        raise RuntimeError(
            "Duplicate year-period combinations."
        )

    if df[
        "real_disposable_income_yoy"
    ].isna().any():

        raise RuntimeError(
            "Missing disposable income values."
        )

    bad = df[
        ~df[
            "real_disposable_income_yoy"
        ].between(
            50,
            150,
        )
    ]

    if not bad.empty:

        raise RuntimeError(
            "Disposable-income values outside "
            "expected range:\n"
            + bad.to_string(
                index=False
            )
        )

    annual = df[
        df["period"] == "year"
    ].copy()

    if annual.empty:

        raise RuntimeError(
            "No annual disposable-income values."
        )

    years = sorted(
        annual["year"]
        .astype(int)
        .unique()
    )

    if years:

        expected = set(
            range(
                min(years),
                max(years) + 1,
            )
        )

        missing = sorted(
            expected
            - set(years)
        )

        if missing:

            raise RuntimeError(
                "Missing annual years: "
                + ", ".join(
                    str(year)
                    for year in missing
                )
            )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "=" * 72
    )

    print(
        "Russia real disposable income updater — Sheet 1"
    )

    print(
        "=" * 72
    )

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    url = (
        find_current_rosstat_file()
    )

    print(
        "\nDownloading current Rosstat workbook..."
    )

    content = download_excel(
        url
    )

    current = (
        parse_real_disposable_income(
            content
        )
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                (
                    "Russia real disposable "
                    "money income"
                ),
                (
                    f"rosstat_page="
                    f"{ROSSTAT_PAGE}"
                ),
                (
                    f"current_file="
                    f"{url}"
                ),
                (
                    f"latest_year="
                    f"{result['year'].max()}"
                ),
                "",
                (
                    "indicator="
                    "real disposable money income"
                ),
                (
                    "unit="
                    "percent to corresponding "
                    "period of previous year"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nRange:",
        result["year"].min(),
        "->",
        result["year"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(15)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()