from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_01_gdp_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_credit_impulse.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_credit_impulse_source.txt"
)


# ============================================================
# CBR
# ============================================================

CBR_BULLETINS_PAGE = (
    "https://www.cbr.ru/ec_research/mb/"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# TEXT HELPERS
# ============================================================

def clean_text(value) -> str:

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def as_number(value) -> float | None:

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):

        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = text.replace(",", ".")

    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:

        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:

        print(
            "WARNING: SSL verification failed."
        )

        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND LATEST STATISTICAL XLSX
# ============================================================

def find_latest_bulletin_xlsx() -> str:

    print("=" * 72)
    print("CBR credit impulse updater")
    print("=" * 72)

    print(
        "\nOpening CBR macroeconomic bulletins page:"
    )

    print(
        CBR_BULLETINS_PAGE
    )

    with requests.Session() as session:

        response = request(
            session,
            "GET",
            CBR_BULLETINS_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    candidates = []

    for tag in soup.find_all(
        "a",
        href=True,
    ):

        href = urljoin(
            CBR_BULLETINS_PAGE,
            tag["href"],
        )

        if not href.lower().split("?")[0].endswith(
            ".xlsx"
        ):
            continue

        text = clean_text(
            tag.get_text(
                " ",
                strip=True,
            )
        )

        parent_text = ""

        parent = tag.parent

        for _ in range(5):

            if parent is None:
                break

            parent_text += " " + clean_text(
                parent.get_text(
                    " ",
                    strip=True,
                )
            )

            parent = parent.parent

        surrounding_text = (
            text
            + " "
            + parent_text
        )

        # We are interested in the statistical annexes
        # to "О чем говорят тренды".
        if (
            "статист" in surrounding_text
            or
            "bulletin_" in href.lower()
        ):
            candidates.append(
                href
            )

    candidates = list(
        dict.fromkeys(
            candidates
        )
    )

    if not candidates:

        raise RuntimeError(
            "No CBR bulletin XLSX candidates found."
        )

    print(
        "\nXLSX candidates found:",
        len(candidates),
    )

    # --------------------------------------------------------
    # Test candidates and choose the newest workbook
    # that actually contains the credit-impulse table.
    # Page order is usually newest first, but we validate
    # the contents rather than trusting the filename.
    # --------------------------------------------------------

    for url in candidates:

        print(
            "\nChecking:"
        )

        print(
            url
        )

        try:

            content = download_xlsx(
                url
            )

            if workbook_has_credit_impulse(
                content
            ):

                print(
                    "\nSelected CBR statistical workbook:"
                )

                print(
                    url
                )

                return url

        except Exception as exc:

            print(
                "Skipped:",
                exc,
            )

    raise RuntimeError(
        "XLSX files were found, but none contained "
        "the credit impulse table."
    )


# ============================================================
# DOWNLOAD
# ============================================================

def download_xlsx(
    url: str,
) -> bytes:

    with requests.Session() as session:

        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    # XLSX = ZIP container
    if not content.startswith(
        b"PK"
    ):
        raise RuntimeError(
            "Downloaded file does not look like XLSX."
        )

    return content


# ============================================================
# FIND CREDIT IMPULSE WORKSHEET
# ============================================================

def find_credit_impulse_sheet(
    wb,
):

    target_phrase = (
        "кредитный импульс, % от ввп"
    )

    for ws in wb.worksheets:

        max_rows = min(
            ws.max_row,
            20,
        )

        max_cols = min(
            ws.max_column,
            10,
        )

        for row in ws.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):

            for value in row:

                text = clean_text(
                    value
                )

                if (
                    target_phrase in text
                    or
                    (
                        "кредитный импульс" in text
                        and
                        "ввп" in text
                    )
                ):

                    return ws

    return None


def workbook_has_credit_impulse(
    content: bytes,
) -> bool:

    wb = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    return (
        find_credit_impulse_sheet(
            wb
        )
        is not None
    )


# ============================================================
# PARSE DATE
# ============================================================

RU_MONTHS = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "сент": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def parse_bulletin_date(
    value,
) -> pd.Timestamp | None:

    if value is None:
        return None

    # Excel date
    if isinstance(
        value,
        pd.Timestamp,
    ):
        return value

    # Python datetime/date
    if hasattr(
        value,
        "year",
    ) and hasattr(
        value,
        "month",
    ):

        try:
            return pd.Timestamp(
                year=value.year,
                month=value.month,
                day=1,
            )

        except Exception:
            pass

    text = clean_text(
        value
    )

    if not text:
        return None

    # Example from the workbook:
    # Jan.15
    # Feb.15
    # Mar.16

    english_match = re.match(
        r"([a-z]{3})\.?\s*(\d{2,4})",
        text,
        flags=re.IGNORECASE,
    )

    if english_match:

        month_text = (
            english_match
            .group(1)
            .lower()
        )

        year_text = (
            english_match
            .group(2)
        )

        english_months = {
            "jan": 1,
            "feb": 2,
            "mar": 3,
            "apr": 4,
            "may": 5,
            "jun": 6,
            "jul": 7,
            "aug": 8,
            "sep": 9,
            "oct": 10,
            "nov": 11,
            "dec": 12,
        }

        month = english_months.get(
            month_text
        )

        if month is not None:

            year = int(
                year_text
            )

            if year < 100:
                year += 2000

            return pd.Timestamp(
                year=year,
                month=month,
                day=1,
            )

    # Russian month text fallback
    for month_name, month in RU_MONTHS.items():

        if text.startswith(
            month_name
        ):

            year_match = re.search(
                r"(\d{2,4})",
                text,
            )

            if year_match:

                year = int(
                    year_match.group(1)
                )

                if year < 100:
                    year += 2000

                return pd.Timestamp(
                    year=year,
                    month=month,
                    day=1,
                )

    # Generic parser fallback
    parsed = pd.to_datetime(
        value,
        errors="coerce",
        dayfirst=True,
    )

    if not pd.isna(parsed):

        return pd.Timestamp(
            year=parsed.year,
            month=parsed.month,
            day=1,
        )

    return None


# ============================================================
# PARSE CREDIT IMPULSE TABLE
# ============================================================

def parse_credit_impulse(
    content: bytes,
) -> pd.DataFrame:

    wb = load_workbook(
        io.BytesIO(content),
        read_only=False,
        data_only=True,
    )

    ws = find_credit_impulse_sheet(
        wb
    )

    if ws is None:

        raise RuntimeError(
            "Credit impulse worksheet was not found."
        )

    print(
        "\nCredit impulse worksheet:"
    )

    print(
        ws.title
    )

    # --------------------------------------------------------
    # Find header row.
    #
    # Current structure:
    # A = date
    # B = nonfinancial corporations
    # C = financial corporations
    # D = unsecured household + auto
    # E = mortgage
    # F = credit impulse
    #
    # But we identify columns from their names.
    # --------------------------------------------------------

    header_row = None

    date_col = None
    nonfinancial_col = None
    financial_col = None
    unsecured_auto_col = None
    mortgage_col = None
    impulse_col = None

    for row_number in range(
        1,
        min(ws.max_row, 20) + 1,
    ):

        row_texts = {}

        for col_number in range(
            1,
            min(ws.max_column, 12) + 1,
        ):

            row_texts[col_number] = clean_text(
                ws.cell(
                    row_number,
                    col_number,
                ).value
            )

        for col_number, text in row_texts.items():

            if text == "дата":
                date_col = col_number

            elif (
                "требования к нефинансовым"
                in text
            ):
                nonfinancial_col = col_number

            elif (
                "требования к финансовым"
                in text
            ):
                financial_col = col_number

            elif (
                "необеспеченные кредиты"
                in text
                and
                "автокредиты"
                in text
            ):
                unsecured_auto_col = col_number

            elif (
                text == "ипотека"
                or
                text.startswith(
                    "ипотека"
                )
            ):
                mortgage_col = col_number

            elif (
                text == "кредитный импульс"
                or
                (
                    "кредитный импульс" in text
                    and
                    "ввп" not in text
                )
            ):
                impulse_col = col_number

        if (
            date_col is not None
            and
            impulse_col is not None
        ):

            header_row = row_number
            break

    if header_row is None:

        raise RuntimeError(
            "Could not identify credit impulse table headers."
        )

    print(
        "Header row:",
        header_row,
    )

    print(
        "Date column:",
        date_col,
    )

    print(
        "Credit impulse column:",
        impulse_col,
    )

    rows = []

    empty_streak = 0

    for row_number in range(
        header_row + 1,
        ws.max_row + 1,
    ):

        raw_date = ws.cell(
            row_number,
            date_col,
        ).value

        d = parse_bulletin_date(
            raw_date
        )

        if d is None:

            empty_streak += 1

            # Stop after a sufficiently long empty/non-date block.
            if empty_streak >= 10:
                break

            continue

        empty_streak = 0

        impulse = as_number(
            ws.cell(
                row_number,
                impulse_col,
            ).value
        )

        if impulse is None:
            continue

        row = {
            "date": d,
            "credit_impulse": impulse,
            "source": "cbr_dip",
        }

        if nonfinancial_col is not None:

            row[
                "nonfinancial_corporations"
            ] = as_number(
                ws.cell(
                    row_number,
                    nonfinancial_col,
                ).value
            )

        if financial_col is not None:

            row[
                "financial_corporations"
            ] = as_number(
                ws.cell(
                    row_number,
                    financial_col,
                ).value
            )

        if unsecured_auto_col is not None:

            row[
                "unsecured_household_auto"
            ] = as_number(
                ws.cell(
                    row_number,
                    unsecured_auto_col,
                ).value
            )

        if mortgage_col is not None:

            row[
                "mortgage"
            ] = as_number(
                ws.cell(
                    row_number,
                    mortgage_col,
                ).value
            )

        rows.append(
            row
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        raise RuntimeError(
            "No credit impulse observations extracted."
        )

    df["date"] = pd.to_datetime(
        df["date"]
    )

    numeric_columns = [
        column
        for column in [
            "nonfinancial_corporations",
            "financial_corporations",
            "unsecured_household_auto",
            "mortgage",
            "credit_impulse",
        ]
        if column in df.columns
    ]

    for column in numeric_columns:

        df[column] = (
            pd.to_numeric(
                df[column],
                errors="coerce",
            )
            .round(2)
        )

    df = (
        df
        .sort_values(
            "date"
        )
        .drop_duplicates(
            "date",
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    # CSV dates as YYYY-MM-DD
    df["date"] = (
        df["date"]
        .dt.strftime(
            "%Y-%m-%d"
        )
    )

    return df


# ============================================================
# PRESERVE EXISTING HISTORY
# ============================================================

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():
        return current

    print(
        "\nLoading existing credit impulse history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required = {
        "date",
        "credit_impulse",
    }

    if not required.issubset(
        existing.columns
    ):

        raise RuntimeError(
            "Existing credit impulse CSV "
            "has unexpected columns."
        )

    if "source" not in existing.columns:

        existing[
            "source"
        ] = "existing_history"

    current_dates = set(
        current["date"].astype(str)
    )

    history_only = existing[
        ~existing["date"]
        .astype(str)
        .isin(
            current_dates
        )
    ].copy()

    history_only[
        "source"
    ] = "existing_history"

    result = pd.concat(
        [
            history_only,
            current,
        ],
        ignore_index=True,
        sort=False,
    )

    result[
        "_date"
    ] = pd.to_datetime(
        result["date"],
        errors="coerce",
    )

    result = (
        result
        .sort_values(
            "_date"
        )
        .drop_duplicates(
            "date",
            keep="last",
        )
        .drop(
            columns="_date"
        )
        .reset_index(
            drop=True
        )
    )

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:

        raise RuntimeError(
            "Credit impulse output is empty."
        )

    if df["date"].duplicated().any():

        raise RuntimeError(
            "Duplicate dates in credit impulse output."
        )

    if df[
        "credit_impulse"
    ].isna().any():

        raise RuntimeError(
            "Missing credit impulse values."
        )

    bad = df[
        ~df[
            "credit_impulse"
        ].between(
            -30,
            30,
        )
    ]

    if not bad.empty:

        raise RuntimeError(
            "Credit impulse values outside expected range:\n"
            + bad.to_string(
                index=False
            )
        )

    dates = pd.to_datetime(
        df["date"]
    )

    if dates.min().year > 2015:

        raise RuntimeError(
            "Credit impulse history unexpectedly "
            "starts after 2015."
        )

    # Check monthly continuity.
    expected = pd.date_range(
        dates.min(),
        dates.max(),
        freq="MS",
    )

    missing = expected.difference(
        pd.DatetimeIndex(
            dates
        )
    )

    if len(missing) > 0:

        print(
            "\nWARNING: missing months:"
        )

        print(
            [
                x.strftime("%Y-%m")
                for x in missing
            ]
        )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook_url = (
        find_latest_bulletin_xlsx()
    )

    print(
        "\nDownloading selected workbook..."
    )

    content = download_xlsx(
        workbook_url
    )

    current = parse_credit_impulse(
        content
    )

    print(
        "\nCurrent CBR credit impulse data:"
    )

    print(
        current
        .tail(15)
        .to_string(
            index=False
        )
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia credit impulse",
                "source=Bank of Russia",
                "calculation=CBR Research and Forecasting Department (DIP)",
                f"bulletins_page={CBR_BULLETINS_PAGE}",
                f"current_workbook={workbook_url}",
                "indicator=Credit impulse",
                "unit=percent of GDP",
                (
                    "components="
                    "nonfinancial corporations; "
                    "financial corporations; "
                    "unsecured household loans and auto loans; "
                    "mortgage"
                ),
                (
                    f"latest_date="
                    f"{result['date'].max()}"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nRange:"
    )

    print(
        result["date"].min(),
        "->",
        result["date"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(15)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()