from __future__ import annotations

import io
import re
import zipfile
import warnings
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_05_income_unemployment_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_household_debt_service_ratio.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_household_debt_service_ratio_source.txt"
)


# ============================================================
# CBR
# ============================================================

CBR_PAGE = "https://cbr.ru/finstab/review/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def url_filename(url: str) -> str:
    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


def as_number(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)
    text = text.replace(",", ".")

    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:
        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:
        print(
            "WARNING: SSL verification failed."
        )

        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND LATEST STATISTICAL FILE
# ============================================================

def find_latest_statistics_file() -> str:

    print("=" * 72)
    print("Russia household debt service ratio updater")
    print("=" * 72)

    print("\nOpening CBR financial stability page:")
    print(CBR_PAGE)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            CBR_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    wanted_heading = (
        "статистические данные "
        "к обзору финансовой стабильности"
    )

    heading_node = None

    for tag in soup.find_all(True):

        text = clean_text(
            tag.get_text(
                " ",
                strip=True,
            )
        )

        if text == wanted_heading:
            heading_node = tag
            break

    if heading_node is None:
        raise RuntimeError(
            "Could not find the CBR section "
            "'Статистические данные к Обзору финансовой стабильности'."
        )

    print(
        "\nStatistics section found."
    )

    # Search in the section following the heading.
    # We want the FIRST current statistics link,
    # because CBR lists issues from newest to oldest.
    current = heading_node

    for _ in range(20):

        current = current.find_next()

        if current is None:
            break

        if current.name != "a":
            continue

        href_raw = current.get(
            "href"
        )

        if not href_raw:
            continue

        href = urljoin(
            CBR_PAGE,
            href_raw,
        )

        filename = url_filename(
            href
        )

        if filename.endswith(
            (
                ".zip",
                ".xlsx",
                ".xls",
            )
        ):

            print(
                "\nSelected latest statistical file:"
            )

            print(
                href
            )

            return href

    raise RuntimeError(
        "Could not find ZIP/XLSX/XLS "
        "under the CBR statistical data section."
    )


# ============================================================
# DOWNLOAD
# ============================================================

def download_file(
    url: str,
) -> bytes:

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    if len(content) < 500:
        raise RuntimeError(
            "Downloaded CBR statistical file "
            "is unexpectedly small."
        )

    return content


# ============================================================
# EXTRACT EXCEL FROM ZIP OR DIRECT FILE
# ============================================================

def extract_excel_files(
    url: str,
    content: bytes,
) -> list[tuple[str, bytes]]:

    filename = url_filename(
        url
    )

    if filename.endswith(
        ".xlsx"
    ):
        if not content.startswith(
            b"PK"
        ):
            raise RuntimeError(
                "Direct CBR XLSX does not look like XLSX."
            )

        return [
            (
                filename,
                content,
            )
        ]

    if filename.endswith(
        ".xls"
    ):
        return [
            (
                filename,
                content,
            )
        ]

    if filename.endswith(
        ".zip"
    ):

        if not zipfile.is_zipfile(
            io.BytesIO(content)
        ):
            raise RuntimeError(
                "CBR response does not look like ZIP."
            )

        files = []

        with zipfile.ZipFile(
            io.BytesIO(content)
        ) as archive:

            for name in archive.namelist():

                lower = name.lower()

                if not lower.endswith(
                    (
                        ".xlsx",
                        ".xls",
                    )
                ):
                    continue

                files.append(
                    (
                        name,
                        archive.read(
                            name
                        ),
                    )
                )

        if not files:
            raise RuntimeError(
                "ZIP contains no Excel files."
            )

        print(
            "\nExcel files inside archive:",
            len(files),
        )

        for name, _ in files:
            print(
                " -",
                name,
            )

        return files

    raise RuntimeError(
        "Unsupported CBR file format: "
        + filename
    )


# ============================================================
# FIND WORKBOOK WITH DSR
# ============================================================

def find_dsr_workbook(
    excel_files: list[tuple[str, bytes]],
) -> tuple[str, bytes]:

    wanted_phrases = [
        "коэффициент обслуживания долга",
        "кредитам физических лиц",
    ]

    for filename, content in excel_files:

        if not filename.lower().endswith(
            ".xlsx"
        ):
            continue

        try:
            wb = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )

        except Exception:
            continue

        for ws in wb.worksheets:

            max_rows = min(
                ws.max_row,
                20,
            )

            max_cols = min(
                ws.max_column,
                20,
            )

            text_parts = []

            for row in ws.iter_rows(
                min_row=1,
                max_row=max_rows,
                min_col=1,
                max_col=max_cols,
                values_only=True,
            ):

                for value in row:

                    if value is not None:
                        text_parts.append(
                            clean_text(value)
                        )

            sheet_text = " ".join(
                text_parts
            )

            if all(
                phrase in sheet_text
                for phrase in wanted_phrases
            ):

                print(
                    "\nDSR workbook found:"
                )

                print(
                    filename
                )

                print(
                    "Sheet:",
                    ws.title,
                )

                return (
                    filename,
                    content,
                )

    raise RuntimeError(
        "Could not find workbook containing "
        "household debt service ratio."
    )

def as_date(value):
    if value is None:
        return None

    # Уже настоящая дата/datetime
    if hasattr(value, "year") and hasattr(value, "month"):
        try:
            return pd.Timestamp(value)
        except Exception:
            pass

    # Excel serial date
    if isinstance(value, (int, float)):
        try:
            # Excel epoch
            return pd.Timestamp("1899-12-30") + pd.to_timedelta(
                float(value),
                unit="D",
            )
        except Exception:
            pass

    # Текстовая дата
    text = str(value).strip()

    if not text:
        return None

    # 01.01.2019
    match = re.search(
        r"\b(\d{1,2})[./-](\d{1,2})[./-](20\d{2})\b",
        text,
    )

    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))

        try:
            return pd.Timestamp(
                year=year,
                month=month,
                day=day,
            )
        except ValueError:
            return None

    return None

# ============================================================
# PARSE DSR
# ============================================================

def parse_household_dsr(
    content: bytes,
) -> pd.DataFrame:

    wb = load_workbook(
        io.BytesIO(content),
        read_only=False,
        data_only=True,
    )

    target_ws = None

    # First try known/current sheet title.
    for ws in wb.worksheets:

        if clean_text(
            ws.title
        ) == "рис20":

            target_ws = ws
            break

    # Fallback by visible title text.
    if target_ws is None:

        for ws in wb.worksheets:

            text_parts = []

            for row in range(
                1,
                min(ws.max_row, 20) + 1,
            ):

                for col in range(
                    1,
                    min(ws.max_column, 20) + 1,
                ):

                    value = ws.cell(
                        row,
                        col,
                    ).value

                    if value is not None:
                        text_parts.append(
                            clean_text(value)
                        )

            text = " ".join(
                text_parts
            )

            if (
                "коэффициент обслуживания долга" in text
                and
                "кредитам физических лиц" in text
            ):

                target_ws = ws
                break

    if target_ws is None:
        raise RuntimeError(
            "Could not identify DSR worksheet."
        )

    ws = target_ws

    print(
        "\nParsing DSR sheet:",
        ws.title,
    )

    # --------------------------------------------------------
    # Find row with dates
    # --------------------------------------------------------

    date_row = None
    best_date_count = 0

    for row in range(
            1,
            min(ws.max_row, 30) + 1,
    ):

        count = 0

        for col in range(
                1,
                ws.max_column + 1,
        ):

            value = ws.cell(
                row,
                col,
            ).value

            parsed_date = as_date(
                value
            )

            if parsed_date is not None:
                count += 1

        print(
            f"Date candidate row {row}: "
            f"{count} dates"
        )

        if count > best_date_count:
            best_date_count = count
            date_row = row

    if (
            date_row is None
            or best_date_count < 4
    ):
        raise RuntimeError(
            "Could not detect DSR date row."
        )

    print(
        "Selected DSR date row:",
        date_row,
    )

    print(
        "Dates detected:",
        best_date_count,
    )

    print(
        "Date row:",
        date_row,
    )

    # --------------------------------------------------------
    # Find total row
    # --------------------------------------------------------

    total_row = None

    for row in range(
        1,
        min(ws.max_row, 30) + 1,
    ):

        row_text = " ".join(
            clean_text(
                ws.cell(
                    row,
                    col,
                ).value
            )
            for col in range(
                1,
                min(ws.max_column, 8) + 1,
            )
        )

        if (
            "общий итог" in row_text
            or
            "всего" in row_text
        ):

            total_row = row
            break

    if total_row is None:
        raise RuntimeError(
            "Could not detect DSR total row."
        )

    print(
        "Total DSR row:",
        total_row,
    )

    # --------------------------------------------------------
    # Parse date/value pairs
    # --------------------------------------------------------

    rows = []

    for col in range(
        1,
        ws.max_column + 1,
    ):

        date_value = ws.cell(
            date_row,
            col,
        ).value

        parsed_date = as_date(
            date_value
        )

        if parsed_date is None:
            continue

        value = as_number(
            ws.cell(
                total_row,
                col,
            ).value
        )

        if value is None:
            continue

        rows.append(
            {
                "date": parsed_date.date(),
                "household_debt_service_ratio": value,
                "source": "cbr_financial_stability_review",
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:
        raise RuntimeError(
            "No DSR observations extracted."
        )

    df = (
        df
        .sort_values(
            "date"
        )
        .drop_duplicates(
            "date",
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    df[
        "household_debt_service_ratio"
    ] = pd.to_numeric(
        df[
            "household_debt_service_ratio"
        ],
        errors="coerce",
    ).round(2)

    return df

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():
        return current

    print(
        "\nLoading existing DSR history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required = {
        "date",
        "household_debt_service_ratio",
    }

    if not required.issubset(
        existing.columns
    ):
        raise RuntimeError(
            "Existing DSR CSV has unexpected columns."
        )

    existing["date"] = pd.to_datetime(
        existing["date"]
    )

    current["date"] = pd.to_datetime(
        current["date"]
    )

    if "source" not in existing.columns:
        existing[
            "source"
        ] = "existing_history"

    current_dates = set(
        current["date"]
    )

    # Сохраняем только те старые даты,
    # которых больше нет в текущем файле ЦБ.
    history_only = existing[
        ~existing[
            "date"
        ].isin(
            current_dates
        )
    ].copy()

    history_only[
        "source"
    ] = "existing_history"

    result = pd.concat(
        [
            history_only,
            current,
        ],
        ignore_index=True,
        sort=False,
    )

    result = (
        result
        .sort_values(
            "date"
        )
        .drop_duplicates(
            "date",
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    return result

# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:
        raise RuntimeError(
            "Household DSR output is empty."
        )

    if df[
        "date"
    ].duplicated().any():

        raise RuntimeError(
            "Duplicate DSR dates."
        )

    if df[
        "household_debt_service_ratio"
    ].isna().any():

        raise RuntimeError(
            "Missing DSR values."
        )

    bad = df[
        ~df[
            "household_debt_service_ratio"
        ].between(
            0,
            30,
        )
    ]

    if not bad.empty:

        raise RuntimeError(
            "DSR values outside expected range:\n"
            + bad.to_string(
                index=False
            )
        )

    # All points should be quarter starts.
    dates = pd.to_datetime(
        df["date"]
    )

    bad_dates = df[
        ~(
            dates.dt.month.isin(
                [
                    1,
                    4,
                    7,
                    10,
                ]
            )
            &
            dates.dt.day.eq(
                1
            )
        )
    ]

    if not bad_dates.empty:

        raise RuntimeError(
            "Unexpected non-quarter-start dates:\n"
            + bad_dates.to_string(
                index=False
            )
        )

    latest_date = pd.to_datetime(
        df["date"]
    ).max()

    if latest_date < pd.Timestamp(
        "2025-01-01"
    ):

        raise RuntimeError(
            "CBR DSR data look unexpectedly old: "
            + str(
                latest_date.date()
            )
        )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    source_url = (
        find_latest_statistics_file()
    )

    print(
        "\nDownloading CBR statistical data..."
    )

    content = download_file(
        source_url
    )

    excel_files = extract_excel_files(
        source_url,
        content,
    )

    workbook_name, workbook_content = (
        find_dsr_workbook(
            excel_files
        )
    )

    current = parse_household_dsr(
        workbook_content
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia household debt service ratio",
                "source=Bank of Russia",
                f"review_page={CBR_PAGE}",
                f"statistics_file={source_url}",
                f"workbook={workbook_name}",
                (
                    f"latest_date="
                    f"{result['date'].max()}"
                ),
                "",
                (
                    "indicator="
                    "Коэффициент обслуживания долга "
                    "по кредитам физических лиц"
                ),
                "unit=percent",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nRange:",
        result["date"].min(),
        "->",
        result["date"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(12)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()