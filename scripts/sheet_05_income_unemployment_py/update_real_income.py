from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_05_income_unemployment_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_real_income.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_real_income_source.txt"
)


# ============================================================
# ROSSTAT
# ============================================================

ROSSTAT_PAGE = "https://rosstat.gov.ru/folder/13397"

TARGET_TITLE_WORDS = [
    "реальные денежные доходы населения",
    "по субъектам российской федерации",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# TEXT HELPERS
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def url_filename(url: str) -> str:
    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


def as_number(value) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = text.replace(",", ".")

    # Remove footnote marks etc.
    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


def detect_year(value) -> int | None:
    text = clean_text(value)

    match = re.search(
        r"\b(20\d{2})\b",
        text,
    )

    if match is None:
        return None

    year = int(
        match.group(1)
    )

    if 2010 <= year <= 2035:
        return year

    return None


def detect_period(value) -> str | None:
    text = clean_text(value)

    if (
        "i кв" in text
        or "1 кв" in text
        or text in {"i", "1"}
    ):
        return "q1"

    if (
        "ii кв" in text
        or "2 кв" in text
        or text in {"ii", "2"}
    ):
        return "q2"

    if (
        "iii кв" in text
        or "3 кв" in text
        or text in {"iii", "3"}
    ):
        return "q3"

    if (
        "iv кв" in text
        or "4 кв" in text
        or text in {"iv", "4"}
    ):
        return "q4"

    if text == "год":
        return "year"

    return None


def is_russian_federation(value) -> bool:
    if value is None:
        return False

    text = clean_text(value)

    # Убираем сноски, звездочки и прочие символы,
    # но сохраняем буквы и пробелы.
    text = re.sub(
        r"[^а-яёa-z ]",
        "",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text.startswith(
        "российская федерация"
    )


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:
        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:
        print(
            "WARNING: SSL verification failed."
        )
        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND CURRENT ROSSTAT FILE
# ============================================================

def find_current_rosstat_file() -> str:

    print("=" * 72)
    print("Russia real income updater")
    print("=" * 72)

    print("\nOpening Rosstat page:")
    print(ROSSTAT_PAGE)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            ROSSTAT_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    wanted_title = (
        "реальные денежные доходы населения "
        "по субъектам российской федерации"
    )

    # Ищем любой элемент страницы, где есть нужное название.
    title_node = None

    for tag in soup.find_all(True):

        text = clean_text(
            tag.get_text(
                " ",
                strip=True,
            )
        )

        if text == wanted_title:
            title_node = tag
            break

    if title_node is None:

        # Более мягкий fallback.
        for tag in soup.find_all(True):

            text = clean_text(
                tag.get_text(
                    " ",
                    strip=True,
                )
            )

            if (
                "реальные денежные доходы населения" in text
                and
                "по субъектам российской федерации" in text
                and
                "среднедушевые" not in text
                and
                "располагаемые" not in text
            ):
                title_node = tag
                break

    if title_node is None:
        raise RuntimeError(
            "Could not find the Rosstat indicator title "
            "'Реальные денежные доходы населения "
            "по субъектам Российской Федерации'."
        )

    print("\nIndicator title found:")
    print(
        clean_text(
            title_node.get_text(
                " ",
                strip=True,
            )
        )
    )

    # Теперь поднимаемся по родителям и ищем XLSX
    # внутри того же блока показателя.
    container = title_node

    for level in range(8):

        if container is None:
            break

        xlsx_links = []

        for tag in container.find_all(
            "a",
            href=True,
        ):

            href = urljoin(
                ROSSTAT_PAGE,
                tag["href"],
            )

            if url_filename(
                href
            ).endswith(".xlsx"):
                xlsx_links.append(
                    href
                )

        xlsx_links = list(
            dict.fromkeys(
                xlsx_links
            )
        )

        if len(xlsx_links) == 1:

            selected = xlsx_links[0]

            print("\nSelected Rosstat workbook:")
            print(selected)

            return selected

        # Если в контейнере уже много соседних XLSX,
        # значит поднялись слишком высоко.
        if len(xlsx_links) > 1:

            print(
                "Container level",
                level,
                "contains",
                len(xlsx_links),
                "XLSX links; continuing carefully."
            )

        container = container.parent

    raise RuntimeError(
        "Indicator title was found, "
        "but its XLSX link could not be isolated."
    )


# ============================================================
# DOWNLOAD
# ============================================================

def download_xlsx(
    url: str,
) -> bytes:

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    # XLSX is a ZIP file.
    if not content.startswith(
        b"PK"
    ):
        preview = content[:300].decode(
            "utf-8",
            errors="ignore",
        )

        raise RuntimeError(
            "Response does not look like XLSX.\n"
            f"URL: {url}\n"
            f"Preview: {preview}"
        )

    return content


# ============================================================
# WORKBOOK DETECTION
# ============================================================

def workbook_contains_rf_row(
    content: bytes,
) -> bool:

    wb = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    for ws in wb.worksheets:

        max_rows = min(
            ws.max_row,
            30,
        )

        max_cols = min(
            ws.max_column,
            10,
        )

        for row in ws.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            for value in row:

                if is_russian_federation(
                    value
                ):
                    return True

    return False


# ============================================================
# FIND CORRECT SHEET + RF ROW
# ============================================================

def find_sheet_and_rf_row(
    wb,
):
    """
    Find the worksheet containing the
    Russian Federation row.
    """

    for ws in wb.worksheets:

        print(
            "Checking sheet:",
            ws.title,
        )

        for row_number in range(
            1,
            min(ws.max_row, 100) + 1,
        ):

            for column_number in range(
                1,
                min(ws.max_column, 20) + 1,
            ):

                value = ws.cell(
                    row_number,
                    column_number,
                ).value

                if is_russian_federation(
                    value
                ):

                    print(
                        "Selected sheet:",
                        ws.title,
                    )

                    print(
                        "Russian Federation found at:",
                        f"row {row_number}, "
                        f"column {column_number}",
                    )

                    print(
                        "Cell value:",
                        repr(value),
                    )

                    return ws, row_number

    raise RuntimeError(
        "Russian Federation row was not found "
        "in any worksheet."
    )


# ============================================================
# PARSE HEADER STRUCTURE
# ============================================================

def find_year_header_row(
    ws,
    rf_row: int,
) -> int:

    best_row = None
    best_count = 0

    for row_number in range(
        1,
        rf_row,
    ):

        year_count = 0

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            year = detect_year(
                ws.cell(
                    row_number,
                    column_number,
                ).value
            )

            if year is not None:
                year_count += 1

        if year_count > best_count:
            best_count = year_count
            best_row = row_number

    if (
        best_row is None
        or best_count < 2
    ):
        raise RuntimeError(
            "Could not detect year header row."
        )

    print(
        "Year header row:",
        best_row,
    )

    return best_row


def find_period_header_row(
    ws,
    year_row: int,
    rf_row: int,
) -> int:

    best_row = None
    best_count = 0

    for row_number in range(
        year_row + 1,
        rf_row,
    ):

        period_count = 0

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            period = detect_period(
                ws.cell(
                    row_number,
                    column_number,
                ).value
            )

            if period is not None:
                period_count += 1

        if period_count > best_count:
            best_count = period_count
            best_row = row_number

    if (
        best_row is None
        or best_count < 4
    ):
        raise RuntimeError(
            "Could not detect quarter/year header row."
        )

    print(
        "Period header row:",
        best_row,
    )

    return best_row


# ============================================================
# BUILD COLUMN MAP
# ============================================================

def build_column_map(
    ws,
    year_row: int,
    period_row: int,
) -> list[tuple[int, int, str]]:

    """
    Returns:
        [(column_number, year, period), ...]

    Handles merged year headers:
    year written once across 5 columns.
    """

    result = []

    current_year = None

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        year = detect_year(
            ws.cell(
                year_row,
                column_number,
            ).value
        )

        if year is not None:
            current_year = year

        period = detect_period(
            ws.cell(
                period_row,
                column_number,
            ).value
        )

        if (
            current_year is not None
            and period is not None
        ):
            result.append(
                (
                    column_number,
                    current_year,
                    period,
                )
            )

    if not result:
        raise RuntimeError(
            "Could not build year-period column map."
        )

    print(
        "Detected observations in header:",
        len(result),
    )

    print(
        "Detected range:",
        min(x[1] for x in result),
        "->",
        max(x[1] for x in result),
    )

    return result


# ============================================================
# PARSE WORKBOOK
# ============================================================

def parse_real_income(
    content: bytes,
) -> pd.DataFrame:

    print(
        "\nParsing Rosstat workbook..."
    )

    wb = load_workbook(
        io.BytesIO(content),
        read_only=False,
        data_only=True,
    )

    ws, rf_row = (
        find_sheet_and_rf_row(
            wb
        )
    )

    year_row = find_year_header_row(
        ws,
        rf_row,
    )

    period_row = find_period_header_row(
        ws,
        year_row,
        rf_row,
    )

    column_map = build_column_map(
        ws,
        year_row,
        period_row,
    )

    rows = []

    for (
        column_number,
        year,
        period,
    ) in column_map:

        source_value = ws.cell(
            rf_row,
            column_number,
        ).value

        value = as_number(
            source_value
        )

        if value is None:
            continue

        rows.append(
            {
                "year": year,
                "period": period,
                "real_income_yoy": value,
                "source": "rosstat_current",
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:
        raise RuntimeError(
            "No Russia real-income observations "
            "were extracted from Rosstat XLSX."
        )

    period_order = {
        "q1": 1,
        "q2": 2,
        "q3": 3,
        "q4": 4,
        "year": 5,
    }

    df["_period_order"] = (
        df["period"]
        .map(period_order)
    )

    df = (
        df
        .sort_values(
            [
                "year",
                "_period_order",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "period",
            ],
            keep="last",
        )
        .drop(
            columns="_period_order"
        )
        .reset_index(
            drop=True
        )
    )

    df[
        "real_income_yoy"
    ] = (
        pd.to_numeric(
            df[
                "real_income_yoy"
            ],
            errors="coerce",
        )
        .round(2)
    )

    return df


# ============================================================
# PRESERVE EXISTING HISTORY
# ============================================================

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():
        return current

    print(
        "\nLoading existing CSV history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required_columns = {
        "year",
        "period",
        "real_income_yoy",
    }

    if not required_columns.issubset(
        existing.columns
    ):
        raise RuntimeError(
            "Existing russia_real_income.csv "
            "has unexpected columns."
        )

    if "source" not in existing.columns:
        existing["source"] = (
            "existing_history"
        )

    # Current Rosstat publication has priority
    # for periods that it actually contains.
    current_keys = set(
        zip(
            current["year"].astype(int),
            current["period"].astype(str),
        )
    )

    history_only = existing[
        ~existing.apply(
            lambda row: (
                int(row["year"]),
                str(row["period"]),
            ) in current_keys,
            axis=1,
        )
    ].copy()

    result = pd.concat(
        [
            history_only,
            current,
        ],
        ignore_index=True,
        sort=False,
    )

    period_order = {
        "q1": 1,
        "q2": 2,
        "q3": 3,
        "q4": 4,
        "year": 5,
    }

    result["_period_order"] = (
        result["period"]
        .map(period_order)
    )

    result = (
        result
        .sort_values(
            [
                "year",
                "_period_order",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "period",
            ],
            keep="last",
        )
        .drop(
            columns="_period_order"
        )
        .reset_index(
            drop=True
        )
    )

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:
        raise RuntimeError(
            "Real-income output is empty."
        )

    if df.duplicated(
        [
            "year",
            "period",
        ]
    ).any():
        raise RuntimeError(
            "Duplicate year-period combinations."
        )

    if df[
        "real_income_yoy"
    ].isna().any():
        raise RuntimeError(
            "Missing real-income values."
        )

    bad = df[
        ~df[
            "real_income_yoy"
        ].between(
            50,
            150,
        )
    ]

    if not bad.empty:
        raise RuntimeError(
            "Real-income indices outside "
            "expected range:\n"
            + bad.to_string(
                index=False
            )
        )

    periods_allowed = {
        "q1",
        "q2",
        "q3",
        "q4",
        "year",
    }

    bad_periods = set(
        df["period"].unique()
    ) - periods_allowed

    if bad_periods:
        raise RuntimeError(
            "Unexpected periods: "
            + ", ".join(
                sorted(
                    bad_periods
                )
            )
        )

    latest_year = int(
        df["year"].max()
    )

    if latest_year < 2025:
        raise RuntimeError(
            "Rosstat real-income data "
            f"look unexpectedly old: {latest_year}"
        )

    # Check continuity of annual history
    annual = df[
        df["period"] == "year"
    ]

    if not annual.empty:

        years = sorted(
            annual["year"]
            .astype(int)
            .unique()
        )

        expected = set(
            range(
                min(years),
                max(years) + 1,
            )
        )

        missing = sorted(
            expected - set(years)
        )

        if missing:
            print(
                "WARNING: missing annual years:",
                missing,
            )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    file_url = (
        find_current_rosstat_file()
    )

    print(
        "\nDownloading selected workbook..."
    )

    content = download_xlsx(
        file_url
    )

    current = parse_real_income(
        content
    )

    print(
        "\nCurrent Rosstat data:"
    )

    print(
        current.tail(15)
        .to_string(index=False)
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia real money income",
                "source=Rosstat",
                f"page={ROSSTAT_PAGE}",
                f"current_file={file_url}",
                (
                    "indicator="
                    "Real money income of population "
                    "by constituent entities of "
                    "the Russian Federation"
                ),
                (
                    "territory="
                    "Russian Federation"
                ),
                (
                    "unit="
                    "percent to corresponding "
                    "period of previous year"
                ),
                (
                    f"latest_year="
                    f"{result['year'].max()}"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nSource info:"
    )

    print(
        SOURCE_INFO_FILE
    )

    print(
        "\nRange:",
        result["year"].min(),
        "->",
        result["year"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(15)
        .to_string(index=False)
    )


if __name__ == "__main__":
    main()