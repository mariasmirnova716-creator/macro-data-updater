from __future__ import annotations

import io
import re
import warnings
from pathlib import Path

import certifi
import pandas as pd
import requests
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = PROJECT_ROOT / "data" / "sheet_06_exp_imp_data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_FILE = DATA_DIR / "sheet_06_exp_imp.csv"




# ============================================================
# SOURCE
# ============================================================

CBR_PAGE_URL = (
    "https://www.cbr.ru/statistics/"
    "macro_itm/external_sector/pb/"
)

CBR_XLSX_URL = (
    "https://www.cbr.ru/vfs/statistics/"
    "credit_statistics/bop/"
    "bal_of_payments_ap.xlsx"
)

TIMEOUT = 90

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/139.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
}


# ============================================================
# HTTP
# ============================================================

def get_session() -> requests.Session:

    session = requests.Session()
    session.headers.update(HEADERS)

    return session


def download_file(
    session: requests.Session,
    url: str,
) -> bytes:

    try:

        response = session.get(
            url,
            timeout=TIMEOUT,
            verify=certifi.where(),
        )

        response.raise_for_status()

        return response.content

    except requests.exceptions.SSLError:

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.get(
            url,
            timeout=TIMEOUT,
            verify=False,
        )

        response.raise_for_status()

        return response.content


# ============================================================
# TEXT HELPERS
# ============================================================

def clean_text(value: object) -> str:

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip()


def normalize_text(value: object) -> str:

    return clean_text(value).lower()


def parse_number(value: object) -> float | None:

    if value is None:
        return None

    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)

    text = clean_text(value)

    if not text:
        return None

    text = text.replace(" ", "")
    text = text.replace("\xa0", "")
    text = text.replace(",", ".")

    text = re.sub(
        r"[^0-9.\-]",
        "",
        text,
    )

    if not text:
        return None

    try:
        return float(text)

    except ValueError:
        return None


# ============================================================
# PERIOD PARSING
# ============================================================

def parse_quarter_header(
    value: object,
) -> str | None:

    text = normalize_text(value)

    if not text:
        return None

    year_match = re.search(
        r"(20\d{2})",
        text,
    )

    if not year_match:
        return None

    year = int(year_match.group(1))

    quarter = None

    if (
        "1 квартал" in text
        or "i квартал" in text
        or re.search(r"\b1q\b", text)
    ):
        quarter = 1

    elif (
        "2 квартал" in text
        or "ii квартал" in text
        or re.search(r"\b2q\b", text)
    ):
        quarter = 2

    elif (
        "3 квартал" in text
        or "iii квартал" in text
        or re.search(r"\b3q\b", text)
    ):
        quarter = 3

    elif (
        "4 квартал" in text
        or "iv квартал" in text
        or re.search(r"\b4q\b", text)
    ):
        quarter = 4

    if quarter is None:
        return None

    return f"{quarter}Q {year}"


# ============================================================
# WORKBOOK HELPERS
# ============================================================

def find_quarter_sheet(
    workbook,
):

    # Prefer sheet with quarterly data.
    for ws in workbook.worksheets:

        title = normalize_text(ws.title)

        if "кварт" in title:
            return ws

    # Fallback:
    # find a sheet containing several quarter headers.
    for ws in workbook.worksheets:

        hits = 0

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, 15),
            values_only=True,
        ):

            for value in row:

                if parse_quarter_header(value) is not None:
                    hits += 1

        if hits >= 4:
            return ws

    raise RuntimeError(
        "Quarterly sheet was not found in CBR workbook."
    )


def find_header_row(
    ws,
) -> int:

    best_row = None
    best_hits = 0

    for row_idx in range(
        1,
        min(ws.max_row, 20) + 1,
    ):

        hits = 0

        for col_idx in range(
            1,
            ws.max_column + 1,
        ):

            value = ws.cell(
                row=row_idx,
                column=col_idx,
            ).value

            if parse_quarter_header(value) is not None:
                hits += 1

        if hits > best_hits:
            best_hits = hits
            best_row = row_idx

    if best_row is None or best_hits < 4:
        raise RuntimeError(
            "Quarter header row was not found."
        )

    return best_row


def build_period_columns(
    ws,
    header_row: int,
) -> dict[int, str]:

    result: dict[int, str] = {}

    for col_idx in range(
        1,
        ws.max_column + 1,
    ):

        value = ws.cell(
            row=header_row,
            column=col_idx,
        ).value

        period = parse_quarter_header(value)

        if period is not None:
            result[col_idx] = period

    if not result:
        raise RuntimeError(
            "No quarterly columns found."
        )

    return result


# ============================================================
# ROW DETECTION
# ============================================================

def find_label_column(
    ws,
    header_row: int,
) -> int:

    # In the current CBR file the labels are on the left.
    # We detect the column containing "current account".
    for row_idx in range(
        header_row + 1,
        min(ws.max_row, header_row + 40) + 1,
    ):

        for col_idx in range(
            1,
            min(ws.max_column, 10) + 1,
        ):

            text = normalize_text(
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                ).value
            )

            if "счет текущих операций" in text:
                return col_idx

    raise RuntimeError(
        "Indicator label column was not found."
    )


def find_required_rows(
    ws,
    label_col: int,
    header_row: int,
) -> dict[str, int]:

    current_account_row = None

    goods_row = None
    goods_export_row = None
    goods_import_row = None

    services_row = None
    services_export_row = None
    services_import_row = None

    section = None

    for row_idx in range(
        header_row + 1,
        ws.max_row + 1,
    ):

        text = normalize_text(
            ws.cell(
                row=row_idx,
                column=label_col,
            ).value
        )

        if not text:
            continue

        # Current account
        if (
            current_account_row is None
            and "счет текущих операций" in text
        ):
            current_account_row = row_idx
            continue

        # Section: goods
        if text == "товары":
            goods_row = row_idx
            section = "goods"
            continue

        # Section: services
        if text == "услуги":
            services_row = row_idx
            section = "services"
            continue

        # Stop treating later exports/imports as services
        # when another major block starts.
        if text in {
            "оплата труда",
            "инвестиционные доходы",
            "вторичные доходы",
            "счет операций с капиталом",
            "сальдо финансового счета",
        }:
            section = None

        if text == "экспорт":

            if (
                section == "goods"
                and goods_export_row is None
            ):
                goods_export_row = row_idx

            elif (
                section == "services"
                and services_export_row is None
            ):
                services_export_row = row_idx

        elif text == "импорт":

            if (
                section == "goods"
                and goods_import_row is None
            ):
                goods_import_row = row_idx

            elif (
                section == "services"
                and services_import_row is None
            ):
                services_import_row = row_idx

    rows = {
        "current_account": current_account_row,
        "goods_export": goods_export_row,
        "goods_import": goods_import_row,
        "services_export": services_export_row,
        "services_import": services_import_row,
    }

    missing = [
        name
        for name, row in rows.items()
        if row is None
    ]

    if missing:
        raise RuntimeError(
            "Required CBR rows were not found: "
            + ", ".join(missing)
        )

    return {
        key: int(value)
        for key, value in rows.items()
    }


# ============================================================
# EXTRACT
# ============================================================

def extract_bop(
    xlsx_bytes: bytes,
) -> pd.DataFrame:

    workbook = load_workbook(
        io.BytesIO(xlsx_bytes),
        data_only=True,
        read_only=False,
    )

    ws = find_quarter_sheet(workbook)

    print(
        f"Using worksheet: {ws.title}"
    )

    header_row = find_header_row(ws)

    print(
        f"Quarter header row: {header_row}"
    )

    period_columns = build_period_columns(
        ws,
        header_row,
    )

    label_col = find_label_column(
        ws,
        header_row,
    )

    print(
        f"Indicator column: {label_col}"
    )

    rows = find_required_rows(
        ws,
        label_col,
        header_row,
    )

    print("Detected rows:")

    for name, row_idx in rows.items():
        print(
            f"  {name}: row {row_idx}"
        )

    output = []

    for col_idx, period in period_columns.items():

        quarter_match = re.match(
            r"([1-4])Q\s+(20\d{2})",
            period,
        )

        if quarter_match is None:
            continue

        quarter_num = int(quarter_match.group(1))
        year = int(quarter_match.group(2))

        current_account = parse_number(
            ws.cell(
                row=rows["current_account"],
                column=col_idx,
            ).value
        )

        goods_export = parse_number(
            ws.cell(
                row=rows["goods_export"],
                column=col_idx,
            ).value
        )

        goods_import = parse_number(
            ws.cell(
                row=rows["goods_import"],
                column=col_idx,
            ).value
        )

        services_export = parse_number(
            ws.cell(
                row=rows["services_export"],
                column=col_idx,
            ).value
        )

        services_import = parse_number(
            ws.cell(
                row=rows["services_import"],
                column=col_idx,
            ).value
        )

        if all(
            value is None
            for value in [
                current_account,
                goods_export,
                goods_import,
                services_export,
                services_import,
            ]
        ):
            continue

        export_total = None
        import_total = None

        if (
            goods_export is not None
            and services_export is not None
        ):
            export_total = (
                goods_export
                + services_export
            )

        if (
            goods_import is not None
            and services_import is not None
        ):
            import_total = (
                goods_import
                + services_import
            )

        output.append(
            {
                "quarter": period,
                "year": year,
                "quarter_num": quarter_num,
                "export_goods": goods_export,
                "export_services": services_export,
                "export_total": export_total,
                "import_goods": goods_import,
                "import_services": services_import,
                "import_total": import_total,
                "current_account": current_account,
            }
        )

    df = pd.DataFrame(output)

    if df.empty:
        raise RuntimeError(
            "No quarterly BOP observations extracted."
        )

    df = (
        df
        .sort_values(
            ["year", "quarter_num"]
        )
        .drop_duplicates(
            subset=["year", "quarter_num"],
            keep="last",
        )
        .reset_index(drop=True)
    )

    return df


# ============================================================
# VALIDATION
# ============================================================

def validate_output(
    df: pd.DataFrame,
) -> None:
    required = {
        "quarter",
        "year",
        "quarter_num",
        "export_goods",
        "export_services",
        "export_total",
        "import_goods",
        "import_services",
        "import_total",
        "current_account",
    }

    missing = required - set(df.columns)

    if missing:
        raise RuntimeError(
            f"Missing output columns: {sorted(missing)}"
        )

    if df.empty:
        raise RuntimeError(
            "Output is empty."
        )

    if df.duplicated(
            subset=["year", "quarter_num"]
    ).any():
        raise RuntimeError(
            "Duplicate quarters detected."
        )

    df_check = df[
        df["year"] >= 2019
        ]

    if df_check.empty:
        raise RuntimeError(
            "No observations from 2019 onward."
        )




    # Basic arithmetic validation.
    exp_diff = (
        df["export_total"]
        - (
            df["export_goods"]
            + df["export_services"]
        )
    ).abs()

    if exp_diff.dropna().gt(0.01).any():
        raise RuntimeError(
            "Export total arithmetic check failed."
        )

    imp_diff = (
        df["import_total"]
        - (
            df["import_goods"]
            + df["import_services"]
        )
    ).abs()

    if imp_diff.dropna().gt(0.01).any():
        raise RuntimeError(
            "Import total arithmetic check failed."
        )


# ============================================================
# MAIN
# ============================================================

def main() -> None:

    session = get_session()

    print(
        "Downloading latest CBR analytical "
        "balance-of-payments workbook..."
    )

    xlsx_bytes = download_file(
        session,
        CBR_XLSX_URL,
    )

    print(
        f"Downloaded {len(xlsx_bytes):,} bytes."
    )

    df = extract_bop(
        xlsx_bytes
    )

    df = df[
        df["year"] >= 2019
        ].copy()

    validate_output(df)

    df = df[
        [
            "quarter",
            "export_goods",
            "export_services",
            "export_total",
            "import_goods",
            "import_services",
            "import_total",
            "current_account",
        ]
    ].copy()

    df.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8-sig",
    )

    print()
    print("Saved:")
    print(OUTPUT_FILE)

    print()
    print(
        df.tail(12).to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()