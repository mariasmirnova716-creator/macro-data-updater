from __future__ import annotations

import io
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = PROJECT_ROOT / "data" / "sheet_2_ipp_data"

OLD_IND_SUB = DATA_DIR / "ind_sub_2018_12-2025.xlsx"
OLD_SEASONAL = DATA_DIR / "sezon_12-2025.xlsx"
OLD_QUARTERLY = DATA_DIR / "ind_baza_2018_4kv-2025.xlsx"

HISTORY_FILE = DATA_DIR / "russia_ipp_history_2020_2023.csv"

MONTHLY_OUTPUT = DATA_DIR / "russia_ipp_monthly.csv"
QUARTERLY_OUTPUT = DATA_DIR / "russia_ipp_quarterly.csv"
PERIODS_OUTPUT = DATA_DIR / "russia_ipp_periods.csv"

SOURCE_INFO_FILE = DATA_DIR / "russia_ipp_source.txt"

ROSSTAT_PAGE = "https://rosstat.gov.ru/enterprise_industrial"

EXPECTED_BASE_YEAR = 2023


MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.strip().lower()

    text = re.sub(r"\s+", " ", text)

    return text


def as_number(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)

    if text in {"", "-", ".", "-."}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def detect_year(value) -> int | None:
    text = clean_text(value)

    match = re.search(r"(20\d{2})", text)

    if match:
        return int(match.group(1))

    return None


def month_end(year: int, month: int) -> pd.Timestamp:
    return (
        pd.Timestamp(year=year, month=month, day=1)
        + pd.offsets.MonthEnd(0)
    )


def find_rf_row(ws) -> int:
    for r in range(1, ws.max_row + 1):

        value = clean_text(ws.cell(r, 1).value)

        if value == "российская федерация":
            return r

    raise RuntimeError(
        f"Russian Federation row not found on sheet {ws.title!r}."
    )


def require_file(path: Path):
    if not path.exists():
        raise FileNotFoundError(
            f"Required file not found:\n{path}"
        )


def request_bytes(url: str) -> bytes:
    """
    Normal HTTPS first.
    Local macOS Python sometimes has Rosstat certificate issues,
    so retry without certificate verification only after SSL failure.
    GitHub Actions normally uses the verified path.
    """

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 Chrome/126 Safari/537.36"
        )
    }

    try:
        response = requests.get(
            url,
            timeout=90,
            headers=headers,
            verify=certifi.where(),
        )

    except requests.exceptions.SSLError:

        print("WARNING: SSL verification failed.")
        print("Retrying Rosstat request without certificate verification.")

        warnings.simplefilter("ignore", InsecureRequestWarning)

        response = requests.get(
            url,
            timeout=90,
            headers=headers,
            verify=False,
        )

    response.raise_for_status()

    return response.content


def download_xlsx(url: str) -> bytes:
    content = request_bytes(url)

    if not content.startswith(b"PK"):
        raise RuntimeError(
            "Downloaded Rosstat file does not look like XLSX:\n"
            + url
        )

    return content


def url_filename(url: str) -> str:
    return unquote(
        url.split("/")[-1].split("?")[0]
    ).lower()


def period_score(url: str) -> tuple[int, int]:
    """
    Examples:
      sezon_2023_06-2026.xlsx
      ind_sub_2023-06-2026.xlsx
    """

    name = url_filename(url)

    matches = re.findall(
        r"(\d{1,2})[-_](20\d{2})",
        name,
    )

    if matches:
        month, year = matches[-1]
        return int(year), int(month)

    return 0, 0


# ============================================================
# FIND CURRENT ROSSTAT FILES
# ============================================================

def find_current_rosstat_files() -> dict[str, str]:

    print("Opening Rosstat page:")
    print(ROSSTAT_PAGE)

    html = request_bytes(ROSSTAT_PAGE).decode(
        "utf-8",
        errors="ignore",
    )

    soup = BeautifulSoup(html, "html.parser")

    seasonal = []
    ind_sub = []
    quarterly = []

    for tag in soup.find_all("a", href=True):

        href = urljoin(
            ROSSTAT_PAGE,
            tag["href"],
        )

        name = url_filename(href)

        if not name.endswith(".xlsx"):
            continue

        # newest monthly seasonal workbook, base 2023
        if re.search(r"sezon[_-]2023", name):
            seasonal.append(href)

        # newest monthly ind_sub workbook, base 2023
        if re.search(r"ind_sub[_-]2023", name):
            ind_sub.append(href)

        # quarterly workbook with base 2023
        if (
            "ind_baza" in name
            and "2023" in name
            and "kv" in name
        ):
            quarterly.append(href)

    if not seasonal:
        raise RuntimeError(
            "Current base-2023 seasonal XLSX was not found."
        )

    if not ind_sub:
        raise RuntimeError(
            "Current base-2023 ind_sub XLSX was not found."
        )

    if not quarterly:
        raise RuntimeError(
            "Current base-2023 quarterly XLSX was not found."
        )

    seasonal_url = max(
        seasonal,
        key=period_score,
    )

    ind_sub_url = max(
        ind_sub,
        key=period_score,
    )

    # There is normally one current *_kv file.
    # If several appear, newest filename period wins where available.
    quarterly_url = max(
        quarterly,
        key=period_score,
    )

    print("\nCurrent seasonal:")
    print(seasonal_url)

    print("\nCurrent ind_sub:")
    print(ind_sub_url)

    print("\nCurrent quarterly:")
    print(quarterly_url)

    return {
        "seasonal": seasonal_url,
        "ind_sub": ind_sub_url,
        "quarterly": quarterly_url,
    }


# ============================================================
# SEASONAL PARSER
# ============================================================
def parse_seasonal(source) -> pd.DataFrame:
    """
    Parses both generations of Rosstat seasonal workbooks.

    Works with:
      - old seasonal workbook, e.g. sezon_12-2025.xlsx
      - current base-2023 workbook, e.g. sezon_2023_06-2026.xlsx

    Expected data columns:
      A = month
      B = MoM actual
      C = MoM SA
      D = level relative to base-year monthly average, actual
      E = same level, SA

    The year may be stored in a merged/header row and not necessarily
    in column A, so we search the first 6 cells of every row.
    """

    wb = load_workbook(
        source,
        data_only=True,
        read_only=True,
    )

    if "1" not in wb.sheetnames:
        raise RuntimeError(
            f"Seasonal workbook has no sheet '1'. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb["1"]

    rows = []
    current_year = None

    for row in ws.iter_rows(values_only=True):

        # Year can sit anywhere in the first several cells
        # because of merged headers.
        detected_year = None

        for value in row[:6]:
            year = detect_year(value)

            if year is not None and 2020 <= year <= 2035:
                detected_year = year
                break

        if detected_year is not None:
            current_year = detected_year

        month_name = clean_text(
            row[0] if len(row) > 0 else None
        )

        if month_name not in MONTHS:
            continue

        if current_year is None:
            continue

        rows.append(
            {
                "date": month_end(
                    current_year,
                    MONTHS[month_name],
                ),
                "ipp_mom": as_number(
                    row[1] if len(row) > 1 else None
                ),
                "ipp_mom_sa": as_number(
                    row[2] if len(row) > 2 else None
                ),
                "ipp_level": as_number(
                    row[3] if len(row) > 3 else None
                ),
                "ipp_level_sa": as_number(
                    row[4] if len(row) > 4 else None
                ),
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        raise RuntimeError(
            "No monthly rows extracted from seasonal workbook."
        )

    df = (
        df
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )

    return df

def detect_month(value) -> int | None:
    text = clean_text(value)

    # Remove ordinary digits and superscript footnote digits
    text = re.sub(r"[0-9¹²³⁴⁵⁶⁷⁸⁹⁰]+", "", text)

    # Remove trailing punctuation / footnote marks
    text = re.sub(r"[^а-яё]+$", "", text).strip()

    for month_name, month_number in MONTHS.items():
        if text == month_name:
            return month_number

    return None

# ============================================================
# IND_SUB MONTHLY PARSER
# ============================================================
def parse_horizontal_monthly(
    source,
    sheet_name: str,
    value_name: str,
) -> pd.DataFrame:
    """
    Robust parser for Rosstat horizontal monthly tables.

    It does NOT assume fixed header rows.

    Automatically:
      1. finds the row containing month names;
      2. finds the Russian Federation data row;
      3. detects year headers above the month row;
      4. carries the year forward across merged year blocks.

    Used for:
      sheet 1 -> YoY
      sheet 3 -> MoM actual
    """

    wb = load_workbook(
        source,
        data_only=True,
        read_only=True,
    )

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet {sheet_name!r} missing. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # --------------------------------------------------------
    # 1. Find row with month names
    # --------------------------------------------------------

    month_row = None
    best_month_count = 0

    for r in range(1, min(ws.max_row, 25) + 1):

        count = 0

        for c in range(1, ws.max_column + 1):
            month_number = detect_month(
                ws.cell(r, c).value
            )

            if month_number is not None:
                count += 1

        if count > best_month_count:
            best_month_count = count
            month_row = r

    if month_row is None or best_month_count < 3:
        raise RuntimeError(
            f"Could not detect month header row "
            f"on sheet {sheet_name!r}."
        )

    print(
        f"Detected month row on sheet {sheet_name}: "
        f"{month_row} ({best_month_count} month labels)"
    )

    # --------------------------------------------------------
    # 2. Find Russian Federation row
    # --------------------------------------------------------

    rf_row = find_rf_row(ws)

    print(
        f"Detected Russian Federation row "
        f"on sheet {sheet_name}: {rf_row}"
    )

    # --------------------------------------------------------
    # 3. Parse columns
    #
    # For every month column:
    # search upward in the header area for a year.
    #
    # If merged Excel cells mean the year appears only
    # above the first month of a block, carry it forward.
    # --------------------------------------------------------

    rows = []
    current_year = None

    for col in range(1, ws.max_column + 1):

        month_number = detect_month(
            ws.cell(month_row, col).value
        )

        # First inspect all rows ABOVE the month row
        # in this same column for a year.
        detected_year = None

        for header_row in range(
            month_row - 1,
            0,
            -1,
        ):
            year = detect_year(
                ws.cell(header_row, col).value
            )

            if year is not None:
                detected_year = year
                break

        if detected_year is not None:
            current_year = detected_year

        if month_number is None:
            continue

        if current_year is None:
            continue

        value = as_number(
            ws.cell(rf_row, col).value
        )

        if value is None:
            continue

        rows.append(
            {
                "date": month_end(
    current_year,
    month_number,
),
                value_name: value,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        raise RuntimeError(
            f"No monthly values extracted "
            f"from sheet {sheet_name!r}."
        )

    df = (
        df
        .sort_values("date")
        .drop_duplicates(
            "date",
            keep="last",
        )
        .reset_index(drop=True)
    )

    print(
        f"Parsed {value_name}: "
        f"{df['date'].min().date()} -> "
        f"{df['date'].max().date()}, "
        f"{len(df)} rows"
    )

    return df


# ============================================================
# IND_SUB CUMULATIVE PERIODS
# ============================================================

def normalize_period_label(value) -> str:
    text = clean_text(value)

    text = text.replace("–", "-")
    text = text.replace("—", "-")

    return text


def cumulative_period_code(label: str) -> str | None:
    """
    We need only rows that actually exist in the main Excel:
      Jan-Mar
      Jan-Jun
      Jan-Sep
      Jan-Dec / year
    """

    label = normalize_period_label(label)

    mappings = {
        "январь-март": "q1_cumulative",
        "январь-июнь": "h1",
        "январь-сентябрь": "m9",
        "январь-декабрь": "year",
    }

    return mappings.get(label)


def parse_ind_sub_periods(source) -> pd.DataFrame:
    """
    Sheet 2 of ind_sub:
    percent to corresponding period of previous year.

    Output:
      year
      period
      ipp_yoy_period
    """

    wb = load_workbook(
        source,
        data_only=True,
        read_only=True,
    )

    if "2" not in wb.sheetnames:
        raise RuntimeError(
            "ind_sub workbook has no sheet '2'."
        )

    ws = wb["2"]

    rf_row = find_rf_row(ws)

    rows = []
    current_year = None

    for col in range(2, ws.max_column + 1):

        detected = detect_year(
            ws.cell(4, col).value
        )

        if detected is not None:
            current_year = detected

        label = normalize_period_label(
            ws.cell(5, col).value
        )

        period = cumulative_period_code(label)

        if period is None:
            continue

        if current_year is None:
            continue

        value = as_number(
            ws.cell(rf_row, col).value
        )

        if value is None:
            continue

        rows.append(
            {
                "year": current_year,
                "period": period,
                "ipp_yoy_period": value,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        raise RuntimeError(
            "No cumulative periods extracted from ind_sub sheet 2."
        )

    return (
        df
        .sort_values(["year", "period"])
        .drop_duplicates(
            ["year", "period"],
            keep="last",
        )
        .reset_index(drop=True)
    )


# ============================================================
# QUARTERLY PARSER
# ============================================================

def quarter_number(value) -> int | None:

    text = clean_text(value)

    match = re.search(
        r"\b([1-4])\s*кварт",
        text,
    )

    if match:
        return int(match.group(1))

    return None


def parse_quarterly_sheet(
    source,
    sheet_name: str,
    value_name: str,
) -> pd.DataFrame:
    """
    Quarterly ind_baza files.

    Sheet 1:
      percent to corresponding quarter previous year

    Sheet 2:
      percent to previous quarter
    """

    wb = load_workbook(
        source,
        data_only=True,
        read_only=True,
    )

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Quarterly workbook has no sheet {sheet_name!r}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    rf_row = None

    for r in range(1, ws.max_row + 1):
        code = clean_text(ws.cell(r, 2).value)

        if code == "bcde":
            rf_row = r
            break

    if rf_row is None:
        raise RuntimeError(
            f"Industrial production row with code BCDE not found on sheet {ws.title!r}."
        )

    rows = []
    current_year = None

    for col in range(3, ws.max_column + 1):

        detected = detect_year(
            ws.cell(4, col).value
        )

        if detected is not None:
            current_year = detected

        q = quarter_number(
            ws.cell(5, col).value
        )

        if q is None:
            continue

        if current_year is None:
            continue

        value = as_number(
            ws.cell(rf_row, col).value
        )

        if value is None:
            continue

        rows.append(
            {
                "year": current_year,
                "quarter": q,
                value_name: value,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        raise RuntimeError(
            f"No quarterly rows extracted from sheet {sheet_name}."
        )

    return (
        df
        .sort_values(["year", "quarter"])
        .drop_duplicates(
            ["year", "quarter"],
            keep="last",
        )
        .reset_index(drop=True)
    )


def parse_quarterly(source) -> pd.DataFrame:

    qoq = parse_quarterly_sheet(
        source,
        sheet_name="1",
        value_name="ipp_qoq",
    )

    yoy = parse_quarterly_sheet(
        source,
        sheet_name="2",
        value_name="ipp_qoy",
    )

    result = yoy.merge(
        qoq,
        on=["year", "quarter"],
        how="outer",
        validate="one_to_one",
    )

    # We currently have no reproducible official quarterly SA source.
    result["ipp_qoq_sa"] = pd.NA

    return (
        result
        .sort_values(["year", "quarter"])
        .reset_index(drop=True)
    )

# ============================================================
# NEWEST-WINS MERGE
# ============================================================

def newest_wins(
    older: pd.DataFrame,
    newer: pd.DataFrame,
    key: str,
) -> pd.DataFrame:
    """
    Newer values replace older values on overlapping dates.
    Older data remain as fallback where newer data are absent.
    """

    result = pd.concat(
        [older, newer],
        ignore_index=True,
        sort=False,
    )

    return (
        result
        .sort_values(key)
        .drop_duplicates(key, keep="last")
        .reset_index(drop=True)
    )


# ============================================================
# MONTHLY BUILD
# ============================================================

def build_monthly(
    current_ind_content: bytes,
    current_season_content: bytes,
) -> pd.DataFrame:

    print("\nParsing old ind_sub monthly YoY...")
    old_yoy = parse_horizontal_monthly(
        OLD_IND_SUB,
        "1",
        "ipp_yoy",
    )

    print("Parsing old ind_sub monthly MoM...")
    old_mom = parse_horizontal_monthly(
        OLD_IND_SUB,
        "3",
        "ipp_mom",
    )

    print("Parsing current ind_sub monthly YoY...")
    current_yoy = parse_horizontal_monthly(
        io.BytesIO(current_ind_content),
        "1",
        "ipp_yoy",
    )

    print("Parsing old seasonal workbook...")
    old_season = parse_seasonal(
        OLD_SEASONAL
    )

    print("Parsing current seasonal workbook...")
    current_season = parse_seasonal(
        io.BytesIO(current_season_content)
    )

    # --------------------------------------------------------
    # YoY
    #
    # Old ind_sub is authoritative through its final month.
    # Current base-2023 ind_sub supplies dates after that.
    # --------------------------------------------------------

    old_yoy_last = old_yoy["date"].max()

    yoy_future = current_yoy[
        current_yoy["date"] > old_yoy_last
    ].copy()

    yoy = pd.concat(
        [old_yoy, yoy_future],
        ignore_index=True,
    )

    yoy = (
        yoy
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )

    yoy["source_yoy"] = yoy["date"].apply(
        lambda d:
        "ind_sub_2018_12-2025"
        if d <= old_yoy_last
        else "ind_sub_2023_current"
    )

    # --------------------------------------------------------
    # MoM actual
    #
    # Through 2025 use long old ind_sub.
    # New seasonal supplies only dates after old ind_sub ends.
    # --------------------------------------------------------

    old_mom_last = old_mom["date"].max()

    current_mom_future = current_season[
        current_season["date"] > old_mom_last
    ][["date", "ipp_mom"]].copy()

    mom = pd.concat(
        [
            old_mom,
            current_mom_future,
        ],
        ignore_index=True,
    )

    mom = (
        mom
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )

    mom["source_mom"] = mom["date"].apply(
        lambda d:
        "ind_sub_2018_12-2025"
        if d <= old_mom_last
        else "sezon_2023_current"
    )

    # --------------------------------------------------------
    # MoM SA
    #
    # IMPORTANT:
    # 2020-2022 are intentionally NOT automated.
    #
    # From 2023:
    # newest seasonal wins.
    # Old seasonal is fallback only for dates missing
    # in current seasonal.
    # --------------------------------------------------------

    old_sa = old_season[
        old_season["date"].dt.year >= 2023
    ][["date", "ipp_mom_sa"]].copy()

    old_sa["source_mom_sa"] = "sezon_12-2025"

    new_sa = current_season[
        current_season["date"].dt.year >= 2023
    ][["date", "ipp_mom_sa"]].copy()

    new_sa["source_mom_sa"] = "sezon_2023_current"

    mom_sa = newest_wins(
        old_sa,
        new_sa,
        key="date",
    )

    # --------------------------------------------------------
    # Levels relative to monthly average 2023
    #
    # 2020-2023:
    # preserve reconstructed values already created in seed.
    #
    # 2024+:
    # newest current seasonal official values.
    # --------------------------------------------------------

    history = pd.read_csv(
        HISTORY_FILE,
        parse_dates=["date"],
    )

    required_history = [
        "date",
        "ipp_level_2023",
        "ipp_level_2023_sa",
    ]

    missing = [
        c
        for c in required_history
        if c not in history.columns
    ]

    if missing:
        raise RuntimeError(
            f"History seed lacks columns: {missing}"
        )

    history_levels = history[
        history["date"].dt.year <= 2023
    ][
        [
            "date",
            "ipp_level_2023",
            "ipp_level_2023_sa",
        ]
    ].copy()

    history_levels["level_reconstructed"] = True
    history_levels["source_level"] = "history_seed_reconstructed"
    history_levels["source_level_sa"] = "history_seed_reconstructed"

    current_levels = current_season[
        current_season["date"].dt.year >= 2024
    ][
        [
            "date",
            "ipp_level",
            "ipp_level_sa",
        ]
    ].copy()

    current_levels = current_levels.rename(
        columns={
            "ipp_level": "ipp_level_2023",
            "ipp_level_sa": "ipp_level_2023_sa",
        }
    )

    current_levels["level_reconstructed"] = False
    current_levels["source_level"] = "sezon_2023_current"
    current_levels["source_level_sa"] = "sezon_2023_current"

    levels = pd.concat(
        [
            history_levels,
            current_levels,
        ],
        ignore_index=True,
    )

    levels = (
        levels
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )

    # --------------------------------------------------------
    # Master date range
    # --------------------------------------------------------

    latest_date = max(
        yoy["date"].max(),
        mom["date"].max(),
        mom_sa["date"].max(),
        levels["date"].max(),
    )

    dates = pd.DataFrame(
        {
            "date": pd.date_range(
                start="2020-01-31",
                end=latest_date,
                freq="ME",
            )
        }
    )

    result = dates.merge(
        yoy,
        on="date",
        how="left",
    )

    result = result.merge(
        mom,
        on="date",
        how="left",
    )

    result = result.merge(
        mom_sa,
        on="date",
        how="left",
    )

    result = result.merge(
        levels,
        on="date",
        how="left",
    )

    result["base_year"] = EXPECTED_BASE_YEAR

    # Explicitly guarantee that automated SA does not exist
    # for 2020-2022.
    legacy_sa_mask = result["date"].dt.year <= 2022

    result.loc[
        legacy_sa_mask,
        "ipp_mom_sa",
    ] = pd.NA

    result.loc[
        legacy_sa_mask,
        "source_mom_sa",
    ] = "legacy_excel_keep_existing"

    # Round source data to Rosstat precision.
    for col in [
        "ipp_yoy",
        "ipp_mom",
        "ipp_mom_sa",
    ]:
        result[col] = pd.to_numeric(
            result[col],
            errors="coerce",
        ).round(1)

    # Keep reconstructed levels with additional precision.
    official_mask = result["level_reconstructed"].eq(False)

    result.loc[
        official_mask,
        "ipp_level_2023",
    ] = pd.to_numeric(
        result.loc[
            official_mask,
            "ipp_level_2023",
        ],
        errors="coerce",
    ).round(1)

    result.loc[
        official_mask,
        "ipp_level_2023_sa",
    ] = pd.to_numeric(
        result.loc[
            official_mask,
            "ipp_level_2023_sa",
        ],
        errors="coerce",
    ).round(1)

    reconstructed_mask = result["level_reconstructed"].eq(True)

    result.loc[
        reconstructed_mask,
        "ipp_level_2023",
    ] = pd.to_numeric(
        result.loc[
            reconstructed_mask,
            "ipp_level_2023",
        ],
        errors="coerce",
    ).round(3)

    result.loc[
        reconstructed_mask,
        "ipp_level_2023_sa",
    ] = pd.to_numeric(
        result.loc[
            reconstructed_mask,
            "ipp_level_2023_sa",
        ],
        errors="coerce",
    ).round(3)

    return result


# ============================================================
# QUARTERLY BUILD
# ============================================================

def build_quarterly(
    current_quarter_content: bytes,
) -> pd.DataFrame:

    print("\nParsing old quarterly workbook...")
    old = parse_quarterly(
        OLD_QUARTERLY
    )

    print("Parsing current quarterly workbook...")
    current = parse_quarterly(
        io.BytesIO(current_quarter_content)
    )

    old_last = (
        old[["year", "quarter"]]
        .sort_values(["year", "quarter"])
        .iloc[-1]
    )

    old_last_key = (
        int(old_last["year"]),
        int(old_last["quarter"]),
    )

    # Old base-2018 quarterly series is used through its
    # final available quarter.
    #
    # Current base-2023 quarterly workbook is used only
    # for later quarters.
    current_future = current[
        current.apply(
            lambda row:
            (int(row["year"]), int(row["quarter"]))
            > old_last_key,
            axis=1,
        )
    ].copy()

    old["source_qoy"] = "ind_baza_2018_4kv-2025"
    old["source_qoq"] = "ind_baza_2018_4kv-2025"

    current_future["source_qoy"] = "ind_baza_2023_current"
    current_future["source_qoq"] = "ind_baza_2023_current"

    result = pd.concat(
        [
            old,
            current_future,
        ],
        ignore_index=True,
        sort=False,
    )

    result = (
        result
        .sort_values(["year", "quarter"])
        .drop_duplicates(
            ["year", "quarter"],
            keep="last",
        )
        .reset_index(drop=True)
    )

    # No official reproducible quarterly SA source.
    result["ipp_qoq_sa"] = pd.NA
    result["source_qoq_sa"] = "legacy_excel_keep_existing"

    for col in [
        "ipp_qoy",
        "ipp_qoq",
    ]:
        result[col] = pd.to_numeric(
            result[col],
            errors="coerce",
        ).round(1)

    return result


# ============================================================
# CUMULATIVE PERIOD BUILD
# ============================================================

def build_periods(
    current_ind_content: bytes,
) -> pd.DataFrame:

    print("\nParsing old cumulative periods...")
    old = parse_ind_sub_periods(
        OLD_IND_SUB
    )

    print("Parsing current cumulative periods...")
    current = parse_ind_sub_periods(
        io.BytesIO(current_ind_content)
    )

    old_last_year = int(
        old["year"].max()
    )

    # Same principle:
    # use the old long workbook through its last year,
    # current workbook after that.
    current_future = current[
        current["year"] > old_last_year
    ].copy()

    old["source_period"] = "ind_sub_2018_12-2025"
    current_future["source_period"] = "ind_sub_2023_current"

    result = pd.concat(
        [
            old,
            current_future,
        ],
        ignore_index=True,
    )

    result = (
        result
        .sort_values(["year", "period"])
        .drop_duplicates(
            ["year", "period"],
            keep="last",
        )
        .reset_index(drop=True)
    )

    result["ipp_yoy_period"] = pd.to_numeric(
        result["ipp_yoy_period"],
        errors="coerce",
    ).round(1)

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate_monthly(df: pd.DataFrame):

    if df.empty:
        raise RuntimeError("Monthly output is empty.")

    if df["date"].duplicated().any():
        raise RuntimeError(
            "Duplicate dates in monthly output."
        )

    expected = pd.date_range(
        df["date"].min(),
        df["date"].max(),
        freq="ME",
    )

    missing_dates = expected.difference(
        pd.DatetimeIndex(df["date"])
    )

    if len(missing_dates):
        raise RuntimeError(
            "Missing months: "
            + ", ".join(
                d.strftime("%Y-%m-%d")
                for d in missing_dates
            )
        )

    if df["date"].min() != pd.Timestamp("2020-01-31"):
        raise RuntimeError(
            "Monthly output does not start in January 2020."
        )

    # YoY and actual MoM should be complete.
    for col in ["ipp_yoy", "ipp_mom"]:

        bad = df[df[col].isna()]

        if not bad.empty:
            raise RuntimeError(
                f"Missing {col}: "
                + ", ".join(
                    bad["date"]
                    .dt.strftime("%Y-%m-%d")
                    .tolist()
                )
            )

    # SA is intentionally allowed to be empty only through 2022.
    sa_required = df[
        df["date"].dt.year >= 2023
    ]

    bad_sa = sa_required[
        sa_required["ipp_mom_sa"].isna()
    ]

    if not bad_sa.empty:
        print(
            "WARNING: no reproducible monthly SA for: "
            + ", ".join(
                bad_sa["date"]
                .dt.strftime("%Y-%m-%d")
                .tolist()
            )
        )

    latest = df["date"].max()
    today = pd.Timestamp.today().normalize()

    if (today - latest).days > 120:
        raise RuntimeError(
            f"Monthly IPP data are too old: {latest.date()}"
        )


def validate_quarterly(df: pd.DataFrame):

    if df.empty:
        raise RuntimeError(
            "Quarterly output is empty."
        )

    if df.duplicated(
        ["year", "quarter"]
    ).any():
        raise RuntimeError(
            "Duplicate year-quarter combinations."
        )

    if df["ipp_qoy"].isna().any():
        raise RuntimeError(
            "Missing quarterly YoY values."
        )

    if df["ipp_qoq"].isna().any():
        raise RuntimeError(
            "Missing quarterly QoQ values."
        )

    # ipp_qoq_sa is allowed to stay empty by design.


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 72)
    print("Russia industrial production updater — Sheet 2")
    print("=" * 72)

    for path in [
        OLD_IND_SUB,
        OLD_SEASONAL,
        OLD_QUARTERLY,
        HISTORY_FILE,
    ]:
        require_file(path)

    urls = find_current_rosstat_files()

    print("\nDownloading current Rosstat workbooks...")

    current_season_content = download_xlsx(
        urls["seasonal"]
    )

    current_ind_content = download_xlsx(
        urls["ind_sub"]
    )

    current_quarter_content = download_xlsx(
        urls["quarterly"]
    )

    monthly = build_monthly(
        current_ind_content=current_ind_content,
        current_season_content=current_season_content,
    )

    quarterly = build_quarterly(
        current_quarter_content=current_quarter_content,
    )

    periods = build_periods(
        current_ind_content=current_ind_content,
    )

    validate_monthly(monthly)
    validate_quarterly(quarterly)

    monthly_columns = [
        "date",
        "ipp_yoy",
        "ipp_mom",
        "ipp_mom_sa",
        "ipp_level_2023",
        "ipp_level_2023_sa",
        "level_reconstructed",
        "base_year",
        "source_yoy",
        "source_mom",
        "source_mom_sa",
        "source_level",
        "source_level_sa",
    ]

    quarterly_columns = [
        "year",
        "quarter",
        "ipp_qoy",
        "ipp_qoq",
        "ipp_qoq_sa",
        "source_qoy",
        "source_qoq",
        "source_qoq_sa",
    ]

    periods_columns = [
        "year",
        "period",
        "ipp_yoy_period",
        "source_period",
    ]

    monthly[
        monthly_columns
    ].to_csv(
        MONTHLY_OUTPUT,
        index=False,
        date_format="%Y-%m-%d",
        encoding="utf-8",
    )

    quarterly[
        quarterly_columns
    ].to_csv(
        QUARTERLY_OUTPUT,
        index=False,
        encoding="utf-8",
    )

    periods[
        periods_columns
    ].to_csv(
        PERIODS_OUTPUT,
        index=False,
        encoding="utf-8",
    )

    retrieved_at = datetime.now(
        timezone.utc
    ).isoformat()

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia industrial production — Sheet 2",
                f"updated_utc={retrieved_at}",
                f"rosstat_page={ROSSTAT_PAGE}",
                "",
                f"current_seasonal={urls['seasonal']}",
                f"current_ind_sub={urls['ind_sub']}",
                f"current_quarterly={urls['quarterly']}",
                "",
                f"old_ind_sub={OLD_IND_SUB.name}",
                f"old_seasonal={OLD_SEASONAL.name}",
                f"old_quarterly={OLD_QUARTERLY.name}",
                f"history_seed={HISTORY_FILE.name}",
                "",
                "monthly_sa_2020_2022=legacy_excel_keep_existing",
                "quarterly_sa=legacy_excel_keep_existing",
                f"base_year={EXPECTED_BASE_YEAR}",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print("\n" + "=" * 72)
    print("DONE")
    print("=" * 72)

    print(f"\nMonthly output:   {MONTHLY_OUTPUT}")
    print(f"Quarterly output: {QUARTERLY_OUTPUT}")
    print(f"Periods output:   {PERIODS_OUTPUT}")
    print(f"Source info:      {SOURCE_INFO_FILE}")

    print(f"\nMonthly rows: {len(monthly)}")
    print(
        "Monthly range:",
        monthly["date"].min().date(),
        "->",
        monthly["date"].max().date(),
    )

    print(f"\nQuarterly rows: {len(quarterly)}")
    print(
        "Quarterly range:",
        f"{quarterly.iloc[0]['year']} Q{quarterly.iloc[0]['quarter']}",
        "->",
        f"{quarterly.iloc[-1]['year']} Q{quarterly.iloc[-1]['quarter']}",
    )

    print("\nLast 18 monthly rows:")
    print(
        monthly[
            monthly_columns
        ].tail(18).to_string(index=False)
    )

    print("\nLast 12 quarterly rows:")
    print(
        quarterly[
            quarterly_columns
        ].tail(12).to_string(index=False)
    )

    print("\nLast cumulative periods:")
    print(
        periods[
            periods_columns
        ].tail(12).to_string(index=False)
    )


if __name__ == "__main__":
    main()