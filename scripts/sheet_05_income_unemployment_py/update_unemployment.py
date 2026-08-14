from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_05_income_unemployment_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_unemployment.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_unemployment_source.txt"
)


# ============================================================
# ROSSTAT
# ============================================================

ROSSTAT_PAGE = "https://rosstat.gov.ru/labour_force"

INDICATOR_TITLE = (
    "Численность безработных в возрасте "
    "15 лет и старше и уровень безработицы"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# TEXT HELPERS
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def url_filename(url: str) -> str:
    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


def as_number(value) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = text.replace(",", ".")

    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


def is_russian_federation(value) -> bool:
    if value is None:
        return False

    text = clean_text(value)

    text = re.sub(
        r"[^а-яёa-z ]",
        "",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text.startswith(
        "российская федерация"
    )


# ============================================================
# CALENDAR QUARTERS
# ============================================================

def detect_calendar_quarter(value):
    if value is None:
        return None

    text = clean_text(value)

    # В Росстате год и номер сноски могут быть склеены:
    # 20171) = 2017 + сноска 1)
    # 20262) = 2026 + сноска 2)
    #
    # Поэтому ничего с цифрами заранее НЕ удаляем.
    # Просто достаем первые четыре цифры года 20xx.
    year_match = re.search(
        r"\b(20\d{2})",
        text,
    )

    if year_match is None:
        return None

    year = int(
        year_match.group(1)
    )

    quarter_patterns = {
        1: ("январь", "март"),
        2: ("апрель", "июнь"),
        3: ("июль", "сентябрь"),
        4: ("октябрь", "декабрь"),
    }

    for quarter, (
        first_month,
        last_month,
    ) in quarter_patterns.items():

        if (
            first_month in text
            and last_month in text
        ):
            return year, quarter

    return None


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:
        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:
        print(
            "WARNING: SSL verification failed."
        )

        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND CURRENT ROSSTAT FILE
# ============================================================

def find_current_rosstat_file() -> str:

    print("=" * 72)
    print("Russia unemployment updater")
    print("=" * 72)

    print("\nOpening Rosstat page:")
    print(ROSSTAT_PAGE)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            ROSSTAT_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    wanted_title = clean_text(
        INDICATOR_TITLE
    )

    title_node = None

    # First: exact visible title.
    for tag in soup.find_all(True):

        text = clean_text(
            tag.get_text(
                " ",
                strip=True,
            )
        )

        if text == wanted_title:
            title_node = tag
            break

    # Fallback: partial title match.
    if title_node is None:

        for tag in soup.find_all(True):

            text = clean_text(
                tag.get_text(
                    " ",
                    strip=True,
                )
            )

            if (
                "численность безработных" in text
                and
                "15 лет и старше" in text
                and
                "уровень безработицы" in text
                and
                "15 - 72" not in text
                and
                "15-72" not in text
            ):
                title_node = tag
                break

    if title_node is None:
        raise RuntimeError(
            "Could not find Rosstat indicator title:\n"
            + INDICATOR_TITLE
        )

    print("\nIndicator title found:")
    print(
        clean_text(
            title_node.get_text(
                " ",
                strip=True,
            )
        )
    )

    # Search XLSX link inside the same indicator card/container.
    container = title_node

    for level in range(8):

        if container is None:
            break

        xlsx_links = []

        for tag in container.find_all(
            "a",
            href=True,
        ):

            href = urljoin(
                ROSSTAT_PAGE,
                tag["href"],
            )

            if url_filename(
                href
            ).endswith(".xlsx"):

                xlsx_links.append(
                    href
                )

        xlsx_links = list(
            dict.fromkeys(
                xlsx_links
            )
        )

        if len(xlsx_links) == 1:

            selected = xlsx_links[0]

            print("\nSelected Rosstat workbook:")
            print(selected)

            return selected

        container = container.parent

    raise RuntimeError(
        "Indicator title was found, "
        "but its XLSX link could not be isolated."
    )


# ============================================================
# DOWNLOAD
# ============================================================

def download_xlsx(
    url: str,
) -> bytes:

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    if not content.startswith(
        b"PK"
    ):

        preview = content[:300].decode(
            "utf-8",
            errors="ignore",
        )

        raise RuntimeError(
            "Response does not look like XLSX.\n"
            f"URL: {url}\n"
            f"Preview: {preview}"
        )

    return content


# ============================================================
# FIND DATA SHEET + RF ROW
# ============================================================

def find_sheet_and_rf_row(
    wb,
):
    """
    Rosstat workbook structure:

    Sheet 4 =
    'Уровень безработицы населения в возрасте
    15 лет и старше по субъектам Российской Федерации,
    в среднем за три месяца'

    This is the series required for the database.
    """

    TARGET_SHEET = "4"

    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"Required Rosstat worksheet '{TARGET_SHEET}' "
            f"was not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[TARGET_SHEET]

    print(
        "Selected sheet:",
        ws.title,
    )

    # Find Russian Federation row only on sheet 4.
    for row_number in range(
        1,
        min(ws.max_row, 150) + 1,
    ):

        for column_number in range(
            1,
            min(ws.max_column, 20) + 1,
        ):

            value = ws.cell(
                row_number,
                column_number,
            ).value

            if is_russian_federation(
                value
            ):

                print(
                    "Russian Federation found at:",
                    f"row {row_number}, "
                    f"column {column_number}",
                )

                print(
                    "Cell value:",
                    repr(value),
                )

                return (
                    ws,
                    row_number,
                )

    raise RuntimeError(
        "Russian Federation row was not found "
        "on Rosstat worksheet 4."
    )


# ============================================================
# FIND HEADER ROW
# ============================================================

def find_period_header_row(
    ws,
    rf_row: int,
) -> int:

    # В таблице Росстата периоды находятся
    # непосредственно над строкой "Российская Федерация".
    candidate_row = rf_row - 1

    print("\n===== REAL VALUES OF HEADER ROW =====")

    for col in range(1, ws.max_column + 1):
        value = ws.cell(candidate_row, col).value

        if value is not None:
            print(
                f"R{candidate_row}C{col}:",
                repr(value)
            )

    print("===== END HEADER VALUES =====\n")

    detected_count = 0

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        value = ws.cell(
            candidate_row,
            column_number,
        ).value

        detected = detect_calendar_quarter(
            value
        )

        if detected is not None:
            detected_count += 1

    print(
        "Period header candidate row:",
        candidate_row,
    )

    print(
        "Calendar quarters detected:",
        detected_count,
    )

    # Если Росстат когда-нибудь изменит структуру файла,
    # не молча берем неправильную строку, а ищем выше.
    if detected_count > 0:
        return candidate_row

    print(
        "Direct row above RF did not work. "
        "Searching previous rows..."
    )

    for row_number in range(
        rf_row - 2,
        0,
        -1,
    ):

        detected_count = 0

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            value = ws.cell(
                row_number,
                column_number,
            ).value

            if detect_calendar_quarter(
                value
            ) is not None:
                detected_count += 1

        if detected_count > 0:

            print(
                "Period header row:",
                row_number,
            )

            print(
                "Calendar quarters detected:",
                detected_count,
            )

            return row_number

    raise RuntimeError(
        "Could not detect unemployment period header row."
    )


# ============================================================
# PARSE UNEMPLOYMENT
# ============================================================

def parse_unemployment(
    content: bytes,
) -> pd.DataFrame:

    print(
        "\nParsing Rosstat workbook..."
    )

    wb = load_workbook(
        io.BytesIO(content),
        read_only=False,
        data_only=True,
    )

    ws, rf_row = (
        find_sheet_and_rf_row(
            wb
        )
    )

    header_row = (
        find_period_header_row(
            ws,
            rf_row,
        )
    )

    rows = []

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        header_value = ws.cell(
            header_row,
            column_number,
        ).value

        detected = detect_calendar_quarter(
            header_value
        )

        if detected is None:
            continue

        year, quarter = detected

        source_value = ws.cell(
            rf_row,
            column_number,
        ).value

        unemployment_rate = as_number(
            source_value
        )

        if unemployment_rate is None:
            continue

        rows.append(
            {
                "year": year,
                "quarter": quarter,
                "unemployment_rate": unemployment_rate,
                "source": "rosstat_current",
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:
        raise RuntimeError(
            "No unemployment observations "
            "were extracted from Rosstat XLSX."
        )

    df = (
        df
        .sort_values(
            [
                "year",
                "quarter",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "quarter",
            ],
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    df[
        "unemployment_rate"
    ] = (
        pd.to_numeric(
            df[
                "unemployment_rate"
            ],
            errors="coerce",
        )
        .round(2)
    )

    return df


# ============================================================
# PRESERVE EXISTING HISTORY
# ============================================================

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():
        return current

    print(
        "\nLoading existing unemployment history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required_columns = {
        "year",
        "quarter",
        "unemployment_rate",
    }

    if not required_columns.issubset(
        existing.columns
    ):
        raise RuntimeError(
            "Existing russia_unemployment.csv "
            "has unexpected columns."
        )

    if "source" not in existing.columns:
        existing[
            "source"
        ] = "existing_history"

    current_keys = set(
        zip(
            current["year"].astype(int),
            current["quarter"].astype(int),
        )
    )

    history_only = existing[
        ~existing.apply(
            lambda row: (
                int(row["year"]),
                int(row["quarter"]),
            ) in current_keys,
            axis=1,
        )
    ].copy()

    result = pd.concat(
        [
            history_only,
            current,
        ],
        ignore_index=True,
        sort=False,
    )

    result = (
        result
        .sort_values(
            [
                "year",
                "quarter",
            ]
        )
        .drop_duplicates(
            [
                "year",
                "quarter",
            ],
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:
        raise RuntimeError(
            "Unemployment output is empty."
        )

    if df.duplicated(
        [
            "year",
            "quarter",
        ]
    ).any():
        raise RuntimeError(
            "Duplicate year-quarter combinations."
        )

    if df[
        "unemployment_rate"
    ].isna().any():
        raise RuntimeError(
            "Missing unemployment-rate values."
        )

    bad_quarters = df[
        ~df[
            "quarter"
        ].isin(
            [
                1,
                2,
                3,
                4,
            ]
        )
    ]

    if not bad_quarters.empty:
        raise RuntimeError(
            "Unexpected quarter numbers:\n"
            + bad_quarters.to_string(
                index=False
            )
        )

    bad_values = df[
        ~df[
            "unemployment_rate"
        ].between(
            0,
            20,
        )
    ]

    if not bad_values.empty:
        raise RuntimeError(
            "Unemployment values outside "
            "expected range:\n"
            + bad_values.to_string(
                index=False
            )
        )

    latest_year = int(
        df["year"].max()
    )

    if latest_year < 2025:
        raise RuntimeError(
            "Rosstat unemployment data "
            f"look unexpectedly old: {latest_year}"
        )

    # Check duplicate calendar quarters inside each year.
    counts = (
        df.groupby(
            [
                "year",
                "quarter",
            ]
        )
        .size()
    )

    if (
        counts > 1
    ).any():
        raise RuntimeError(
            "Duplicate unemployment quarters detected."
        )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    file_url = (
        find_current_rosstat_file()
    )

    print(
        "\nDownloading selected workbook..."
    )

    content = download_xlsx(
        file_url
    )

    current = parse_unemployment(
        content
    )

    print(
        "\nCurrent Rosstat unemployment data:"
    )

    print(
        current.tail(16)
        .to_string(
            index=False
        )
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia unemployment rate",
                "source=Rosstat",
                f"page={ROSSTAT_PAGE}",
                f"current_file={file_url}",
                (
                    "indicator="
                    "Численность безработных "
                    "в возрасте 15 лет и старше "
                    "и уровень безработицы"
                ),
                (
                    "territory="
                    "Российская Федерация"
                ),
                (
                    "frequency="
                    "calendar quarters only"
                ),
                (
                    "quarter_definition="
                    "Q1 Jan-Mar; "
                    "Q2 Apr-Jun; "
                    "Q3 Jul-Sep; "
                    "Q4 Oct-Dec"
                ),
                (
                    "unit="
                    "percent of labor force"
                ),
                (
                    f"latest_year="
                    f"{result['year'].max()}"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nSource info:"
    )

    print(
        SOURCE_INFO_FILE
    )

    print(
        "\nRange:",
        result["year"].min(),
        "->",
        result["year"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(16)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()