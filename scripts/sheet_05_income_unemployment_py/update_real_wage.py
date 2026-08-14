from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin, unquote

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_05_income_unemployment_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_real_wage_annual.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_real_wage_annual_source.txt"
)


# ============================================================
# ROSSTAT
# ============================================================

ROSSTAT_PAGE = (
    "https://rosstat.gov.ru/"
    "labor_market_employment_salaries"
)

INDICATOR_TITLE = (
    "Реальная начисленная заработная плата, "
    "в % к соответствующему периоду предыдущего года"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# TEXT HELPERS
# ============================================================

def clean_text(value) -> str:

    if value is None:
        return ""

    text = str(value)

    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip().lower()


def url_filename(url: str) -> str:

    return unquote(
        url.split("/")[-1]
        .split("?")[0]
    ).lower()


def as_number(value) -> float | None:

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):

        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = text.replace(",", ".")

    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


def is_russian_federation(
    value,
) -> bool:

    text = clean_text(
        value
    )

    text = re.sub(
        r"[^а-яёa-z ]",
        "",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text.startswith(
        "российская федерация"
    )


def detect_year(
    value,
) -> int | None:

    if value is None:
        return None

    text = clean_text(
        value
    )

    # Росстат может хранить год со сноской:
    # 2022¹)
    # 20221)
    # 2023 1)
    #
    # Нам нужны только первые четыре цифры года.
    match = re.search(
        r"(20\d{2})",
        text,
    )

    if match is None:
        return None

    year = int(
        match.group(1)
    )

    if (
        year < 2000
        or year > 2100
    ):
        return None

    return year


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    kwargs.setdefault(
        "timeout",
        90,
    )

    kwargs.setdefault(
        "headers",
        HEADERS,
    )

    try:

        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:

        print(
            "WARNING: SSL verification failed."
        )

        print(
            "Retrying without certificate verification."
        )

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()

    return response


# ============================================================
# FIND CURRENT ROSSTAT FILE
# ============================================================

def find_current_rosstat_file() -> str:

    print("=" * 72)
    print("Russia real wage updater")
    print("=" * 72)

    print("\nOpening Rosstat labour-market page:")
    print(ROSSTAT_PAGE)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            ROSSTAT_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    # ========================================================
    # 1. Find the SMALLEST element containing indicator title
    # ========================================================

    matches = []

    for tag in soup.find_all(True):

        text = clean_text(
            tag.get_text(
                " ",
                strip=True,
            )
        )

        if (
            "реальная среднемесячная начисленная заработная плата"
            in text
            and
            "по субъектам российской федерации"
            in text
        ):

            matches.append(
                (
                    len(text),
                    tag,
                    text,
                )
            )

    if not matches:

        raise RuntimeError(
            "Could not find Rosstat indicator title "
            "for real average accrued wage."
        )

    # The smallest matching HTML element is normally
    # the actual indicator title rather than the whole page.
    matches.sort(
        key=lambda item: item[0]
    )

    _, title_node, title_text = matches[0]

    print("\nIndicator title found:")
    print(title_text)

    # ========================================================
    # 2. Find the first XLSX link AFTER this title
    # ========================================================

    checked_links = 0

    for link in title_node.find_all_next(
        "a",
        href=True,
    ):

        href = urljoin(
            ROSSTAT_PAGE,
            link["href"],
        )

        filename = url_filename(
            href
        )

        # Do not search indefinitely into the entire page.
        checked_links += 1

        if filename.endswith(".xlsx"):

            print("\nXLSX button found:")
            print(href)

            # =================================================
            # 3. Confirm that the workbook really contains
            #    the expected real-wage table
            # =================================================

            print("\nChecking workbook contents...")

            content = download_excel(
                href
            )

            if not workbook_contains_real_wage(
                content
            ):

                raise RuntimeError(
                    "The XLSX located next to the real-wage "
                    "indicator does not contain the expected table."
                )

            print("\nSelected Rosstat workbook:")
            print(href)

            return href

        if checked_links >= 10:
            break

    raise RuntimeError(
        "Real-wage indicator title was found, "
        "but no XLSX button was found immediately after it."
    )


# ============================================================
# DOWNLOAD
# ============================================================

def download_excel(
    url: str,
) -> bytes:

    with requests.Session() as session:

        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    # XLSX = ZIP container.
    if content.startswith(
        b"PK"
    ):
        return content

    raise RuntimeError(
        "Downloaded Rosstat file "
        "does not look like XLSX:\n"
        + url
    )


# ============================================================
# CONFIRM WORKBOOK
# ============================================================

def workbook_contains_real_wage(
    content: bytes,
) -> bool:

    wb = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    for ws in wb.worksheets:

        found_rf = False
        found_years = 0
        found_real_wage_title = False

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(
                ws.max_row,
                20,
            ),
            values_only=True,
        ):

            for value in row:

                text = clean_text(
                    value
                )

                if (
                    "реальная среднемесячная"
                    in text
                    and
                    "заработная плата"
                    in text
                ):
                    found_real_wage_title = True

                if is_russian_federation(
                    value
                ):
                    found_rf = True

                if detect_year(
                    value
                ) is not None:
                    found_years += 1

        if (
            found_rf
            and
            found_years >= 3
            and
            found_real_wage_title
        ):
            return True

    return False


# ============================================================
# FIND DATA SHEET
# ============================================================

def find_data_sheet_and_rf_row(
    wb,
):

    # ========================================================
    # 1. Prefer explicit current-history sheet:
    #    "с 2018"
    # ========================================================

    for ws in wb.worksheets:

        sheet_name = clean_text(
            ws.title
        )

        if (
            "с 2018" in sheet_name
            or
            "с2018" in sheet_name
        ):

            print(
                "Preferred sheet found:",
                ws.title,
            )

            for row_number in range(
                1,
                min(
                    ws.max_row,
                    150,
                ) + 1,
            ):

                for column_number in range(
                    1,
                    min(
                        ws.max_column,
                        20,
                    ) + 1,
                ):

                    value = ws.cell(
                        row_number,
                        column_number,
                    ).value

                    if is_russian_federation(
                        value
                    ):

                        print(
                            "Russian Federation row:",
                            row_number,
                        )

                        return (
                            ws,
                            row_number,
                        )

            raise RuntimeError(
                "Sheet 'с 2018' was found, "
                "but Russian Federation row was not found."
            )

    # ========================================================
    # 2. Fallback:
    #    choose sheet containing RF and the latest year
    # ========================================================

    best_result = None
    best_latest_year = 0

    for ws in wb.worksheets:

        print(
            "Checking fallback sheet:",
            ws.title,
        )

        rf_row = None

        for row_number in range(
            1,
            min(
                ws.max_row,
                150,
            ) + 1,
        ):

            for column_number in range(
                1,
                min(
                    ws.max_column,
                    20,
                ) + 1,
            ):

                value = ws.cell(
                    row_number,
                    column_number,
                ).value

                if is_russian_federation(
                    value
                ):

                    rf_row = row_number
                    break

            if rf_row is not None:
                break

        if rf_row is None:
            continue

        latest_year = 0

        for row_number in range(
            1,
            min(
                rf_row + 2,
                ws.max_row,
            ) + 1,
        ):

            for column_number in range(
                1,
                ws.max_column + 1,
            ):

                year = detect_year(
                    ws.cell(
                        row_number,
                        column_number,
                    ).value
                )

                if (
                    year is not None
                    and
                    year > latest_year
                ):
                    latest_year = year

        print(
            "Latest year on sheet:",
            latest_year,
        )

        if latest_year > best_latest_year:

            best_latest_year = latest_year

            best_result = (
                ws,
                rf_row,
            )

    if best_result is None:

        raise RuntimeError(
            "Could not find Rosstat sheet "
            "containing current real-wage data."
        )

    ws, rf_row = best_result

    print(
        "\nSelected fallback sheet:",
        ws.title,
    )

    print(
        "Russian Federation row:",
        rf_row,
    )

    print(
        "Latest year:",
        best_latest_year,
    )

    return (
        ws,
        rf_row,
    )


# ============================================================
# FIND YEAR HEADER ROW
# ============================================================

def find_year_header_row(
    ws,
    rf_row: int,
) -> int:

    best_row = None
    best_count = 0

    # Usually directly above RF row,
    # but search several preceding rows.
    for row_number in range(
        max(
            1,
            rf_row - 10,
        ),
        rf_row,
    ):

        count = 0

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            year = detect_year(
                ws.cell(
                    row_number,
                    column_number,
                ).value
            )

            if year is not None:
                count += 1

        if count > best_count:

            best_count = count
            best_row = row_number

    if (
        best_row is None
        or
        best_count < 2
    ):

        raise RuntimeError(
            "Could not detect annual year header row."
        )

    print(
        "Year header row:",
        best_row,
    )

    print(
        "Years detected:",
        best_count,
    )

    return best_row


# ============================================================
# PARSE
# ============================================================

def parse_real_wage(
    content: bytes,
) -> pd.DataFrame:

    print(
        "\nParsing Rosstat real-wage workbook..."
    )

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=False,
    )

    ws, rf_row = (
        find_data_sheet_and_rf_row(
            wb
        )
    )

    year_row = (
        find_year_header_row(
            ws,
            rf_row,
        )
    )

    rows = []

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        year = detect_year(
            ws.cell(
                year_row,
                column_number,
            ).value
        )

        if year is None:
            continue

        source_value = ws.cell(
            rf_row,
            column_number,
        ).value

        real_wage_yoy = as_number(
            source_value
        )

        if real_wage_yoy is None:
            continue

        rows.append(
            {
                "year": year,
                "real_wage_yoy": real_wage_yoy,
                "source": "rosstat_current",
            }
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        raise RuntimeError(
            "No annual real-wage observations "
            "were extracted from Rosstat XLSX."
        )

    df = (
        df
        .sort_values(
            "year"
        )
        .drop_duplicates(
            "year",
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    df[
        "real_wage_yoy"
    ] = pd.to_numeric(
        df[
            "real_wage_yoy"
        ],
        errors="coerce",
    ).round(1)

    return df


# ============================================================
# PRESERVE EXISTING HISTORY
# ============================================================

def merge_existing_history(
    current: pd.DataFrame,
) -> pd.DataFrame:

    if not OUTPUT_FILE.exists():

        return current

    print(
        "\nLoading existing real-wage history..."
    )

    existing = pd.read_csv(
        OUTPUT_FILE
    )

    required_columns = {
        "year",
        "real_wage_yoy",
    }

    if not required_columns.issubset(
        existing.columns
    ):

        raise RuntimeError(
            "Existing real-wage CSV "
            "has unexpected columns."
        )

    if "source" not in existing.columns:

        existing[
            "source"
        ] = "existing_history"

    current_years = set(
        current[
            "year"
        ].astype(int)
    )

    # --------------------------------------------------------
    # Preserve historical years that disappeared
    # from the latest Rosstat workbook.
    #
    # Years that ARE present in current Rosstat data
    # are replaced by the latest official values.
    # --------------------------------------------------------

    history_only = existing[
        ~existing[
            "year"
        ].astype(int).isin(
            current_years
        )
    ].copy()

    history_only[
        "source"
    ] = "existing_history"

    result = pd.concat(
        [
            history_only,
            current,
        ],
        ignore_index=True,
        sort=False,
    )

    result = (
        result
        .sort_values(
            "year"
        )
        .drop_duplicates(
            "year",
            keep="last",
        )
        .reset_index(
            drop=True
        )
    )

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:

        raise RuntimeError(
            "Real-wage output is empty."
        )

    if df[
        "year"
    ].duplicated().any():

        raise RuntimeError(
            "Duplicate real-wage years."
        )

    if df[
        "real_wage_yoy"
    ].isna().any():

        raise RuntimeError(
            "Missing real-wage values."
        )

    bad_values = df[
        ~df[
            "real_wage_yoy"
        ].between(
            50,
            150,
        )
    ]

    if not bad_values.empty:

        raise RuntimeError(
            "Real-wage values outside "
            "expected range:\n"
            + bad_values.to_string(
                index=False
            )
        )

    years = sorted(
        df[
            "year"
        ].astype(int)
        .unique()
    )

    if not years:

        raise RuntimeError(
            "No valid real-wage years."
        )

    expected_years = set(
        range(
            min(years),
            max(years) + 1,
        )
    )

    missing_years = sorted(
        expected_years
        - set(years)
    )

    if missing_years:

        raise RuntimeError(
            "Missing real-wage years: "
            + ", ".join(
                str(year)
                for year in missing_years
            )
        )

    latest_year = max(
        years
    )

    if latest_year < 2024:

        raise RuntimeError(
            "Rosstat real-wage data "
            f"look unexpectedly old: {latest_year}"
        )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    file_url = (
        find_current_rosstat_file()
    )

    print(
        "\nDownloading selected Rosstat workbook..."
    )

    content = download_excel(
        file_url
    )

    current = parse_real_wage(
        content
    )

    print(
        "\nCurrent Rosstat real-wage data:"
    )

    print(
        current.tail(12)
        .to_string(
            index=False
        )
    )

    result = merge_existing_history(
        current
    )

    validate(
        result
    )

    result.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia annual real accrued wage",
                "source=Rosstat",
                f"page={ROSSTAT_PAGE}",
                f"current_file={file_url}",
                (
                    "indicator="
                    "Реальная начисленная заработная плата, "
                    "в % к соответствующему периоду "
                    "предыдущего года"
                ),
                "territory=Российская Федерация",
                "frequency=annual",
                "unit=percent to previous year",
                (
                    f"latest_year="
                    f"{result['year'].max()}"
                ),
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nSource info:"
    )

    print(
        SOURCE_INFO_FILE
    )

    print(
        "\nRange:",
        result[
            "year"
        ].min(),
        "->",
        result[
            "year"
        ].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        result
        .tail(12)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()