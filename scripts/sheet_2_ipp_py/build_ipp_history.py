from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# FILES
# ============================================================

OLD_FILE = Path("Чату для листа2.xlsx")
SEASON_FILE = Path("sezon_12-2025.xlsx")
OUTPUT_FILE = Path("data/russia_ipp_history_2020_2023.csv")

MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def detect_year(values) -> int | None:
    for value in values:
        text = clean_text(value)
        match = re.search(r"(20\d{2})", text)
        if match:
            return int(match.group(1))
    return None


def month_end(year: int, month: int) -> pd.Timestamp:
    return pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)


# ============================================================
# 1. OLD SHEET: 2020–2023
# ============================================================

def read_old_sheet() -> pd.DataFrame:
    wb = load_workbook(OLD_FILE, data_only=True, read_only=True)
    ws = wb["2 (ИПП)"]

    rows = []
    current_year = None

    for row in ws.iter_rows(values_only=True):
        # Год может стоять в объединённой ячейке, поэтому смотрим первые 2 поля.
        year = detect_year(row[:2])
        if year is not None and 2020 <= year <= 2026:
            current_year = year

        month_name = clean_text(row[0] if len(row) > 0 else None)
        if month_name not in MONTHS or current_year is None:
            continue

        # Старый лист:
        # B = YoY
        # C = MoM факт
        # D = MoM SA
        # E = старый level факт
        # F = старый level SA
        rows.append(
            {
                "date": month_end(current_year, MONTHS[month_name]),
                "ipp_yoy": as_number(row[1]),
                "ipp_mom": as_number(row[2]),
                "ipp_mom_sa": as_number(row[3]),
            }
        )

    df = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
    df = df[df["date"].dt.year.between(2020, 2023)].copy()

    expected = 48
    if len(df) != expected:
        raise RuntimeError(
            f"Из старого листа ожидалось 48 месячных строк за 2020–2023, получено {len(df)}."
        )

    return df


# ============================================================
# 2. SEZON_12-2025: fresher 2023 MoM + SA
# ============================================================

def read_season_2023() -> pd.DataFrame:
    wb = load_workbook(SEASON_FILE, data_only=True, read_only=True)
    ws = wb["1"]

    rows = []
    current_year = None

    for row in ws.iter_rows(values_only=True):
        # В строке года значение может быть в объединённой области.
        year = detect_year(row[:6])
        if year is not None and 2020 <= year <= 2026:
            current_year = year

        month_name = clean_text(row[0] if len(row) > 0 else None)
        if month_name not in MONTHS or current_year != 2023:
            continue

        # Файл sezon_12-2025, блок "Индекс промышленного производства":
        # B = к предыдущему периоду, факт
        # C = к предыдущему периоду, SA
        # D = к среднемесячному значению 2022, факт
        # E = к среднемесячному значению 2022, SA
        rows.append(
            {
                "date": month_end(2023, MONTHS[month_name]),
                "ipp_mom_new": as_number(row[1]),
                "ipp_mom_sa_new": as_number(row[2]),
                "official_level_oldbase": as_number(row[3]),
                "official_level_oldbase_sa": as_number(row[4]),
            }
        )

    df = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)

    if len(df) != 12:
        raise RuntimeError(
            f"Из sezon_12-2025 ожидалось 12 месяцев 2023 года, получено {len(df)}."
        )

    return df


# ============================================================
# 3. RECONSTRUCT LEVEL 2023=100
# ============================================================

def build_level_from_mom(df: pd.DataFrame, mom_col: str) -> pd.Series:
    """
    Строим цепной уровень. Абсолютный старт произвольный:
    после нормировки на среднее 2023 года он сокращается.
    """
    levels = [100.0]

    for i in range(1, len(df)):
        mom = df.iloc[i][mom_col]
        if pd.isna(mom):
            raise RuntimeError(
                f"Пропуск {mom_col} на {df.iloc[i]['date'].date()} — уровень построить нельзя."
            )

        levels.append(levels[-1] * float(mom) / 100.0)

    s = pd.Series(levels, index=df.index)

    mask_2023 = df["date"].dt.year.eq(2023)
    avg_2023 = s.loc[mask_2023].mean()

    return s / avg_2023 * 100.0


def rebase_official_2023(series: pd.Series) -> pd.Series:
    return series / series.mean() * 100.0


# ============================================================
# MAIN
# ============================================================

def main():
    if not OLD_FILE.exists():
        raise FileNotFoundError(f"Не найден файл: {OLD_FILE}")

    if not SEASON_FILE.exists():
        raise FileNotFoundError(f"Не найден файл: {SEASON_FILE}")

    history = read_old_sheet()
    season_2023 = read_season_2023()

    # 2020–2022 оставляем из старого официального листа.
    # Для 2023 заменяем MoM и SA на более свежую оценку Росстата.
    history = history.merge(
        season_2023[["date", "ipp_mom_new", "ipp_mom_sa_new"]],
        on="date",
        how="left",
    )

    is_2023 = history["date"].dt.year.eq(2023)

    history.loc[is_2023, "ipp_mom"] = history.loc[is_2023, "ipp_mom_new"]
    history.loc[is_2023, "ipp_mom_sa"] = history.loc[is_2023, "ipp_mom_sa_new"]

    history = history.drop(columns=["ipp_mom_new", "ipp_mom_sa_new"])

    # Строим единые уровни, где среднемесячное значение 2023 года = 100.
    history["ipp_level_2023"] = build_level_from_mom(history, "ipp_mom")
    history["ipp_level_2023_sa"] = build_level_from_mom(history, "ipp_mom_sa")

    # Сверяем 2023 с официальными уровнями из sezon_12-2025,
    # предварительно перебазировав их с 2022=100 на 2023=100.
    check = season_2023.copy()
    check["official_2023_rebased"] = rebase_official_2023(
        check["official_level_oldbase"]
    )
    check["official_2023_rebased_sa"] = rebase_official_2023(
        check["official_level_oldbase_sa"]
    )

    check = check.merge(
        history[
            ["date", "ipp_level_2023", "ipp_level_2023_sa"]
        ],
        on="date",
        how="left",
    )

    check["diff_fact"] = (
        check["ipp_level_2023"] - check["official_2023_rebased"]
    ).abs()
    check["diff_sa"] = (
        check["ipp_level_2023_sa"] - check["official_2023_rebased_sa"]
    ).abs()

    max_diff_fact = check["diff_fact"].max()
    max_diff_sa = check["diff_sa"].max()

    # Из-за округления исходных месячных индексов допускаем небольшую разницу.
    if max_diff_fact > 0.35:
        raise RuntimeError(
            f"Слишком большая разница при проверке фактического уровня 2023: "
            f"{max_diff_fact:.3f} п.п."
        )

    if max_diff_sa > 0.35:
        raise RuntimeError(
            f"Слишком большая разница при проверке SA-уровня 2023: "
            f"{max_diff_sa:.3f} п.п."
        )

    history["level_reconstructed"] = True
    history["base_year"] = 2023

    history["source_yoy"] = "old_sheet_2"
    history["source_mom"] = "old_sheet_2"
    history["source_mom_sa"] = "old_sheet_2"

    history.loc[is_2023, "source_mom"] = "sezon_12-2025"
    history.loc[is_2023, "source_mom_sa"] = "sezon_12-2025"

    # Для Excel достаточно одного знака после запятой,
    # но в CSV оставляем 3 знака для reconstructed level.
    history["ipp_yoy"] = history["ipp_yoy"].round(1)
    history["ipp_mom"] = history["ipp_mom"].round(1)
    history["ipp_mom_sa"] = history["ipp_mom_sa"].round(1)
    history["ipp_level_2023"] = history["ipp_level_2023"].round(3)
    history["ipp_level_2023_sa"] = history["ipp_level_2023_sa"].round(3)

    output_columns = [
        "date",
        "ipp_yoy",
        "ipp_mom",
        "ipp_mom_sa",
        "ipp_level_2023",
        "ipp_level_2023_sa",
        "level_reconstructed",
        "base_year",
        "source_yoy",
        "source_mom",
        "source_mom_sa",
    ]

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    history[output_columns].to_csv(
        OUTPUT_FILE,
        index=False,
        date_format="%Y-%m-%d",
        encoding="utf-8",
    )

    print(f"Создано: {OUTPUT_FILE}")
    print(f"Строк: {len(history)}")
    print(
        f"Проверка 2023: max diff fact = {max_diff_fact:.3f} п.п.; "
        f"max diff SA = {max_diff_sa:.3f} п.п."
    )
    print()
    print(history[output_columns].tail(15).to_string(index=False))


if __name__ == "__main__":
    main()
