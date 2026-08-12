from __future__ import annotations

import io
import re
import warnings
from pathlib import Path
from urllib.parse import urljoin

import certifi
import pandas as pd
import requests
from bs4 import BeautifulSoup
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = PROJECT_ROOT / "data" / "sheet_01_gdp_data"
OUTPUT_FILE = DATA_DIR / "russia_average_population.csv"
SOURCE_INFO_FILE = DATA_DIR / "russia_average_population_source.txt"


# ============================================================
# EMISS / FEDSTAT
# ============================================================

INDICATOR_ID = 31556

INDICATOR_URL = (
    f"https://www.fedstat.ru/indicator/{INDICATOR_ID}"
)

ROSSTAT_DEMOGRAPHY_PAGE = "https://rosstat.gov.ru/folder/12781"

DOWNLOAD_URL = (
    "https://www.fedstat.ru/indicator/"
    "downloadData.do?format=excel"
)

# These object IDs came from the actual Fedstat export request.
LINE_OBJECT_IDS = ["57831", "58274"]
COLUMN_OBJECT_IDS = ["3"]

# Russian Federation
RF_FILTER = "57831_1688487"

# All population
ALL_POPULATION_FILTER = "58274_1744150"

# Other filters present in the real export request.
EXTRA_FILTERS = [
    "30611_950458",
    "33560_1558883",
]

FILTER_OBJECT_IDS = [
    "0",
    "30611",
    "33560",
]


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:
    """
    Normal HTTPS first.
    On local macOS Python we retry without SSL verification
    only if certificate verification fails.
    """

    kwargs.setdefault("timeout", 90)
    kwargs.setdefault("headers", HEADERS)

    try:
        response = session.request(
            method,
            url,
            verify=certifi.where(),
            **kwargs,
        )

    except requests.exceptions.SSLError:
        print("WARNING: SSL verification failed.")
        print("Retrying without certificate verification.")

        warnings.simplefilter(
            "ignore",
            InsecureRequestWarning,
        )

        response = session.request(
            method,
            url,
            verify=False,
            **kwargs,
        )

    response.raise_for_status()
    return response


# ============================================================
# PAGE PARSING
# ============================================================

def extract_struts_token(html: str) -> tuple[str, str]:
    """
    Fedstat export POST uses:

        struts.token.name=token
        token=<dynamic value>

    We try several variants because the page markup may differ.
    """

    soup = BeautifulSoup(html, "html.parser")

    token_name_input = soup.find(
        "input",
        attrs={"name": "struts.token.name"},
    )

    token_name = "token"

    if token_name_input is not None:
        value = token_name_input.get("value")

        if value:
            token_name = value

    token_input = soup.find(
        "input",
        attrs={"name": token_name},
    )

    if token_input is not None:
        token_value = token_input.get("value")

        if token_value:
            return token_name, token_value

    # Fallback: search HTML directly.
    match = re.search(
        r'name=["\']token["\']'
        r'[^>]*value=["\']([^"\']+)["\']',
        html,
        flags=re.IGNORECASE,
    )

    if match:
        return "token", match.group(1)

    raise RuntimeError(
        "Could not find Fedstat Struts token."
    )


def extract_available_years(html: str) -> list[int]:
    """
    Find years available on the indicator page.

    We intentionally restrict to plausible statistical years,
    otherwise JS/version numbers may be picked up.
    """

    years = {
        int(x)
        for x in re.findall(r"\b(19\d{2}|20\d{2})\b", html)
    }

    years = {
        year
        for year in years
        if 1990 <= year <= 2035
    }

    if not years:
        raise RuntimeError(
            "Could not detect any years on Fedstat indicator page."
        )

    return sorted(years)


def extract_title(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    text = soup.get_text(
        " ",
        strip=True,
    )

    wanted = (
        "Численность постоянного населения "
        "в среднем за год"
    )

    if wanted.lower() in text.lower():
        return (
            "Численность постоянного населения "
            "в среднем за год "
            "(человек, значение показателя за год)"
        )

    # Exact Russian title is not technically required
    # for the export, but Fedstat included it in the real request.
    return wanted

def rosstat_filename(url: str) -> str:
    return url.split("/")[-1].split("?")[0].lower()


def find_latest_rosstat_population_file() -> tuple[str, int]:
    """
    Find the newest federal Rosstat population estimate workbook.

    Federal workbooks use the OkPopul_CompYYYY_Site.xlsx naming pattern.
    The year in the filename is the population date year;
    average annual population relates to the previous year.
    """

    print("\nOpening Rosstat demography page:")
    print(ROSSTAT_DEMOGRAPHY_PAGE)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            ROSSTAT_DEMOGRAPHY_PAGE,
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser",
    )

    candidates = []

    for tag in soup.find_all("a", href=True):

        href = urljoin(
            ROSSTAT_DEMOGRAPHY_PAGE,
            tag["href"],
        )

        filename = rosstat_filename(href)

        match = re.fullmatch(
            r"okpopul_comp(20\d{2})_site\.xlsx",
            filename,
            flags=re.IGNORECASE,
        )

        if match is None:
            continue

        publication_year = int(
            match.group(1)
        )

        average_year = (
            publication_year - 1
        )

        candidates.append(
            (
                average_year,
                href,
            )
        )

    if not candidates:
        raise RuntimeError(
            "Could not find an OkPopul_CompYYYY_Site.xlsx "
            "workbook on Rosstat demography page."
        )

    average_year, file_url = max(
        candidates,
        key=lambda item: item[0],
    )

    print(
        "Latest Rosstat average-population year:",
        average_year,
    )

    print(
        "Latest Rosstat workbook:"
    )
    print(file_url)

    return file_url, average_year

# ============================================================
# DOWNLOAD XLS
# ============================================================

def download_rosstat_xlsx(url: str) -> bytes:

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            url,
        )

    content = response.content

    if not content.startswith(b"PK"):
        preview = content[:300].decode(
            "utf-8",
            errors="ignore",
        )

        raise RuntimeError(
            "Rosstat response does not look like XLSX.\n"
            f"URL: {url}\n"
            f"Preview: {preview}"
        )

    return content

def download_population_xls() -> tuple[bytes, list[int]]:
    print("=" * 72)
    print("Fedstat average population updater")
    print("=" * 72)

    session = requests.Session()
    session.headers.update(HEADERS)

    print("\nOpening indicator:")
    print(INDICATOR_URL)

    response = request(
        session,
        "GET",
        INDICATOR_URL,
    )

    html = response.text

    print(
        "Session cookies:",
        list(session.cookies.keys()),
    )

    token_name, token_value = extract_struts_token(html)

    print("Token detected.")
    print("Token name:", token_name)

    years = extract_available_years(html)

    print(
        "Detected year range:",
        min(years),
        "->",
        max(years),
    )

    selected_filters = [
        f"0_{INDICATOR_ID}",
    ]

    selected_filters.extend(
        f"3_{year}"
        for year in years
    )

    selected_filters.extend(EXTRA_FILTERS)
    selected_filters.append(RF_FILTER)
    selected_filters.append(ALL_POPULATION_FILTER)

    payload = [
        (
            "title",
            extract_title(html),
        ),
        (
            "struts.token.name",
            token_name,
        ),
        (
            token_name,
            token_value,
        ),
        (
            "id",
            str(INDICATOR_ID),
        ),
    ]

    for object_id in LINE_OBJECT_IDS:
        payload.append(
            ("lineObjectIds", object_id)
        )

    for object_id in COLUMN_OBJECT_IDS:
        payload.append(
            ("columnObjectIds", object_id)
        )

    for filter_id in selected_filters:
        payload.append(
            ("selectedFilterIds", filter_id)
        )

    for object_id in FILTER_OBJECT_IDS:
        payload.append(
            ("filterObjectIds", object_id)
        )

    export_headers = {
        **HEADERS,
        "Accept": (
            "text/html,application/xhtml+xml,"
            "application/xml;q=0.9,*/*;q=0.8"
        ),
        "Content-Type": (
            "application/x-www-form-urlencoded"
        ),
        "Origin": "https://www.fedstat.ru",
        "Referer": INDICATOR_URL,
    }

    print("\nRequesting Excel export...")

    export_response = request(
        session,
        "POST",
        DOWNLOAD_URL,
        headers=export_headers,
        data=payload,
    )

    content = export_response.content

    if len(content) < 1000:
        preview = content[:500].decode(
            "utf-8",
            errors="ignore",
        )

        raise RuntimeError(
            "Fedstat returned an unexpectedly small response.\n"
            f"Preview:\n{preview}"
        )

    print(
        "Downloaded:",
        len(content),
        "bytes",
    )

    return content, years


# ============================================================
# XLS PARSING
# ============================================================

def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def as_int(value) -> int | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None

        return int(round(float(value)))

    text = clean_text(value)

    text = (
        text
        .replace(" ", "")
        .replace("\xa0", "")
    )

    text = re.sub(
        r"[^\d\-]",
        "",
        text,
    )

    if text in {"", "-"}:
        return None

    try:
        return int(text)

    except ValueError:
        return None


def parse_population_xls(content: bytes) -> pd.DataFrame:
    """
    Expected Fedstat layout from the actual downloaded file:

        row 1: title
        row 3: year headers from column C onward
        row 4:
            A = Российская Федерация
            B = все население
            C... = values

    Parser is made a little more flexible so we do not depend
    entirely on exact Excel row numbers.
    """

    excel = pd.ExcelFile(
        io.BytesIO(content),
        engine="xlrd",
    )

    if not excel.sheet_names:
        raise RuntimeError(
            "Downloaded XLS contains no sheets."
        )

    sheet_name = (
        "Данные"
        if "Данные" in excel.sheet_names
        else excel.sheet_names[0]
    )

    raw = pd.read_excel(
        io.BytesIO(content),
        sheet_name=sheet_name,
        header=None,
        engine="xlrd",
    )

    if raw.empty:
        raise RuntimeError(
            "Fedstat population XLS is empty."
        )

    # --------------------------------------------------------
    # Find Russian Federation / all population row
    # --------------------------------------------------------

    rf_row = None

    for idx in raw.index:
        first = clean_text(
            raw.iloc[idx, 0]
            if raw.shape[1] > 0
            else None
        ).lower()

        second = clean_text(
            raw.iloc[idx, 1]
            if raw.shape[1] > 1
            else None
        ).lower()

        if (
            first == "российская федерация"
            and second == "все население"
        ):
            rf_row = idx
            break

    if rf_row is None:
        raise RuntimeError(
            "Row 'Российская Федерация / все население' "
            "was not found in Fedstat XLS."
        )

    # --------------------------------------------------------
    # Find year header row above RF row
    # --------------------------------------------------------

    year_row = None
    best_year_count = 0

    for idx in range(0, rf_row):
        count = 0

        for value in raw.iloc[idx]:
            year = as_int(value)

            if year is not None and 1990 <= year <= 2035:
                count += 1

        if count > best_year_count:
            best_year_count = count
            year_row = idx

    if year_row is None or best_year_count < 3:
        raise RuntimeError(
            "Could not detect Fedstat year header row."
        )

    print(
        "Detected year header row:",
        year_row + 1,
    )

    print(
        "Detected RF data row:",
        rf_row + 1,
    )

    rows = []

    for col in range(raw.shape[1]):
        year = as_int(
            raw.iloc[year_row, col]
        )

        if year is None:
            continue

        if not (1990 <= year <= 2035):
            continue

        population = as_int(
            raw.iloc[rf_row, col]
        )

        if population is None:
            continue

        rows.append(
            {
                "year": year,
                "average_population": population,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        raise RuntimeError(
            "No population values extracted from Fedstat XLS."
        )

    df = (
        df
        .sort_values("year")
        .drop_duplicates(
            "year",
            keep="last",
        )
        .reset_index(drop=True)
    )

    return df


def parse_latest_rosstat_population_xlsx(
    content: bytes,
    expected_year: int,
) -> pd.DataFrame:
    """
    Parse current federal Rosstat population workbook
    without relying on Russian text labels.

    Structural logic:
      - first worksheet is the total-population sheet;
      - RF row contains several values around 100-200 million;
      - average annual population is the rightmost header column
        containing expected_year.
    """

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=False,
    )

    if not wb.worksheets:
        raise RuntimeError(
            "Rosstat workbook contains no worksheets."
        )

    ws = wb.worksheets[0]

    print(
        "Parsing Rosstat sheet:",
        ws.title,
    )

    # ========================================================
    # 1. FIND THE NATIONAL TOTAL ROW
    # ========================================================

    rf_row = None

    for row in range(1, ws.max_row + 1):

        large_values = []

        for col in range(1, ws.max_column + 1):

            value = ws.cell(
                row=row,
                column=col,
            ).value

            number = as_int(value)

            if number is None:
                continue

            if 100_000_000 <= number <= 200_000_000:
                large_values.append(number)

        # In the federal row we currently have, for example:
        # population at 01.01.2024,
        # population at 01.01.2025,
        # average population for 2024.
        if len(large_values) >= 2:
            rf_row = row
            break

    if rf_row is None:
        raise RuntimeError(
            "Could not identify the Russian Federation data row "
            "from population values."
        )

    print(
        "Rosstat national row:",
        rf_row,
    )

    # ========================================================
    # 2. FIND ALL HEADER COLUMNS CONTAINING EXPECTED YEAR
    # ========================================================

    year_columns = []

    for col in range(1, ws.max_column + 1):

        header_parts = []

        for row in range(1, rf_row):

            value = ws.cell(
                row=row,
                column=col,
            ).value

            if value is not None:
                header_parts.append(
                    str(value)
                )

        header_text = " ".join(
            header_parts
        )

        if str(expected_year) in header_text:
            year_columns.append(col)

    if not year_columns:
        raise RuntimeError(
            f"No Rosstat header column contains year "
            f"{expected_year}."
        )

    print(
        "Columns containing year",
        expected_year,
        ":",
        year_columns,
    )

    # In Rosstat's federal workbook the average-annual value
    # is the final/rightmost column referring to that year.
    target_col = max(year_columns)

    print(
        "Selected average-population column:",
        target_col,
    )

    # ========================================================
    # 3. READ VALUE
    # ========================================================

    raw_population = ws.cell(
        row=rf_row,
        column=target_col,
    ).value

    population = as_int(
        raw_population
    )

    if population is None:
        raise RuntimeError(
            "Selected Rosstat population cell "
            "is empty or non-numeric. "
            f"row={rf_row}, col={target_col}, "
            f"value={raw_population!r}"
        )

    if not (
        100_000_000
        <= population
        <= 200_000_000
    ):
        raise RuntimeError(
            "Selected Rosstat population value is outside "
            f"expected range: {population}"
        )

    print(
        "Rosstat average population:",
        expected_year,
        population,
    )

    return pd.DataFrame(
        [
            {
                "year": expected_year,
                "average_population": population,
                "source": "rosstat_current",
            }
        ]
    )
# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
    expected_latest_year: int,
) -> None:

    if df.empty:
        raise RuntimeError(
            "Population output is empty."
        )

    if df["year"].duplicated().any():
        raise RuntimeError(
            "Duplicate years in population output."
        )

    if df["average_population"].isna().any():
        raise RuntimeError(
            "Missing average population values."
        )

    # ------------------------------------------------------------
    # Check continuity of years
    # ------------------------------------------------------------

    first_year = int(
        df["year"].min()
    )

    latest_year = int(
        df["year"].max()
    )

    expected_years = set(
        range(
            first_year,
            latest_year + 1,
        )
    )

    actual_years = set(
        df["year"].astype(int)
    )

    missing_years = sorted(
        expected_years - actual_years
    )

    if missing_years:
        raise RuntimeError(
            "Population history has missing years: "
            + ", ".join(
                str(year)
                for year in missing_years
            )
        )

    bad = df[
        ~df["average_population"].between(
            100_000_000,
            200_000_000,
        )
    ]

    if not bad.empty:
        raise RuntimeError(
            "Population values outside expected range:\n"
            + bad.to_string(index=False)
        )



    if latest_year < expected_latest_year:
        raise RuntimeError(
            "Combined population data are unexpectedly old. "
            f"Expected at least {expected_latest_year}, "
            f"got {latest_year}."
        )


# ============================================================
# MAIN
# ============================================================

def main():
    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    content, page_years = download_population_xls()

    print("\nParsing Fedstat XLS...")

    population = parse_population_xls(
        content
    )

    population["source"] = "fedstat_31556"

    print("\nFinding latest Rosstat population workbook...")

    rosstat_url, rosstat_year = (
        find_latest_rosstat_population_file()
    )

    print("\nDownloading latest Rosstat population workbook...")

    rosstat_content = download_rosstat_xlsx(
        rosstat_url
    )

    print("\nParsing latest Rosstat population workbook...")

    rosstat_current = (
        parse_latest_rosstat_population_xlsx(
            rosstat_content,
            expected_year=rosstat_year,
        )
    )

    # ============================================================
    # PRESERVE EXISTING HISTORY
    # ============================================================

    existing_population = None

    if OUTPUT_FILE.exists():

        print("\nLoading existing population history...")

        existing_population = pd.read_csv(
            OUTPUT_FILE
        )

        required_existing_columns = {
            "year",
            "average_population",
        }

        if not required_existing_columns.issubset(
                existing_population.columns
        ):
            raise RuntimeError(
                "Existing population CSV has unexpected structure."
            )

        if "source" not in existing_population.columns:
            existing_population["source"] = "existing_history"

    # ============================================================
    # MERGE SOURCES
    #
    # Priority:
    # 1. existing CSV
    # 2. current Fedstat
    # 3. current Rosstat
    #
    # Later sources win on overlapping years.
    # ============================================================

    # ============================================================
    # MERGE WITHOUT OVERWRITING EXISTING HISTORY
    #
    # Existing CSV has priority for years already stored.
    # Fedstat and Rosstat are used only to add missing/new years.
    # ============================================================

    if existing_population is not None:

        existing_years = set(
            existing_population["year"]
            .astype(int)
            .tolist()
        )

        fedstat_new = population[
            ~population["year"]
            .astype(int)
            .isin(existing_years)
        ].copy()

        rosstat_new = rosstat_current[
            ~rosstat_current["year"]
            .astype(int)
            .isin(existing_years)
        ].copy()

        population = pd.concat(
            [
                existing_population,
                fedstat_new,
                rosstat_new,
            ],
            ignore_index=True,
            sort=False,
        )

    else:

        population = pd.concat(
            [
                population,
                rosstat_current,
            ],
            ignore_index=True,
            sort=False,
        )

    population = (
        population
        .sort_values("year")
        .drop_duplicates(
            "year",
            keep="first",
        )
        .reset_index(drop=True)
    )

    validate(
        population,
        expected_latest_year=rosstat_year,
    )


    # For our macro database we currently need modern history.
    population = population[
        population["year"] >= 2011
        ].copy()


    population.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia average annual population",
                "source=EMISS / Fedstat",
                f"indicator_id={INDICATOR_ID}",
                f"indicator_url={INDICATOR_URL}",
                f"latest_year={population['year'].max()}",
                (
                    "page_detected_latest_year="
                    f"{max(page_years)}"
                ),
                "",
                (
                    "definition="
                    "Численность постоянного населения "
                    "в среднем за год"
                ),
                "population_type=все население",
                "territory=Российская Федерация",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print("\n" + "=" * 72)
    print("DONE")
    print("=" * 72)

    print("\nOutput:")
    print(OUTPUT_FILE)

    print("\nSource info:")
    print(SOURCE_INFO_FILE)

    print(
        "\nRange:",
        population["year"].min(),
        "->",
        population["year"].max(),
    )

    print("\nLast rows:")
    print(
        population
        .tail(10)
        .to_string(index=False)
    )


if __name__ == "__main__":
    main()