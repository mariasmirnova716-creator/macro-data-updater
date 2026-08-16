from __future__ import annotations

import io
import re
import warnings
from pathlib import Path

import certifi
import pandas as pd
import requests
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning


# ============================================================
# PATHS
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "sheet_01_gdp_data"
)

OUTPUT_FILE = (
    DATA_DIR
    / "russia_extended_debt_quarterly.csv"
)

SOURCE_INFO_FILE = (
    DATA_DIR
    / "russia_extended_debt_quarterly_source.txt"
)


# ============================================================
# SOURCE
# ============================================================

SOURCE_PAGE = (
    "https://www.cbr.ru/statistics/macro_itm/dkfs/"
)

SOURCE_FILE_URL = (
    "https://www.cbr.ru/vfs/statistics/nfs_hd/"
    "debt_extended.xlsx"
)

SOURCE_NAME = "cbr_debt_extended"

FIRST_OUTPUT_DATE = pd.Timestamp("2020-01-01")

TARGET_ROW_TEXT = (
    "Долг организаций и домашних хозяйств, итого"
)


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/148 Safari/537.36"
    ),
    "Accept-Language": "ru,en;q=0.9",
}


# ============================================================
# HTTP
# ============================================================

def request(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs,
) -> requests.Response:

    import time

    kwargs.setdefault("timeout", 90)
    kwargs.setdefault("headers", HEADERS)

    last_error = None

    for attempt in range(1, 5):

        try:

            response = session.request(
                method,
                url,
                verify=certifi.where(),
                **kwargs,
            )

            response.raise_for_status()
            return response

        except requests.exceptions.SSLError as exc:

            last_error = exc

            print(
                f"SSL error, attempt {attempt}/4."
            )

            warnings.simplefilter(
                "ignore",
                InsecureRequestWarning,
            )

            try:

                response = session.request(
                    method,
                    url,
                    verify=False,
                    **kwargs,
                )

                response.raise_for_status()
                return response

            except requests.RequestException as exc2:
                last_error = exc2

        except requests.RequestException as exc:

            last_error = exc

            print(
                f"Connection error, attempt {attempt}/4:"
            )
            print(exc)

        if attempt < 4:
            print("Retrying in 5 seconds...")
            time.sleep(5)

    raise RuntimeError(
        "Could not download CBR file after 4 attempts."
    ) from last_error


# ============================================================
# HELPERS
# ============================================================

def clean_text(value) -> str:

    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip()


def normalize_text(value) -> str:

    return clean_text(value).lower()


def as_number(value) -> float | None:

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):

        if pd.isna(value):
            return None

        return float(value)

    text = clean_text(value)

    text = (
        text
        .replace(" ", "")
        .replace(",", ".")
    )

    if text in {
        "",
        "-",
        "—",
        "…",
        "...",
    }:
        return None

    match = re.search(
        r"-?\d+(?:\.\d+)?",
        text,
    )

    if match is None:
        return None

    try:
        return float(
            match.group(0)
        )

    except ValueError:
        return None


def parse_excel_date(value) -> pd.Timestamp | None:

    if value is None:
        return None

    if hasattr(value, "year") and hasattr(value, "month"):

        try:
            return pd.Timestamp(
                year=int(value.year),
                month=int(value.month),
                day=int(value.day),
            )

        except Exception:
            pass

    text = clean_text(value)

    if not text:
        return None

    for fmt in (
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):

        try:
            return pd.to_datetime(
                text,
                format=fmt,
            ).normalize()

        except Exception:
            pass

    return None


# ============================================================
# DOWNLOAD
# ============================================================

def download_workbook() -> bytes:

    print("=" * 72)
    print("CBR extended debt updater")
    print("=" * 72)

    print("\nDownloading:")
    print(SOURCE_FILE_URL)

    with requests.Session() as session:
        response = request(
            session,
            "GET",
            SOURCE_FILE_URL,
        )

    content = response.content

    if not content.startswith(b"PK"):

        preview = content[:300].decode(
            "utf-8",
            errors="ignore",
        )

        raise RuntimeError(
            "CBR response does not look like XLSX.\n"
            f"Preview:\n{preview}"
        )

    print(
        "Downloaded:",
        len(content),
        "bytes",
    )

    return content


# ============================================================
# FIND TARGET ROW
# ============================================================

def find_target_row(ws) -> int:

    wanted = normalize_text(
        TARGET_ROW_TEXT
    )

    for row in range(
        1,
        ws.max_row + 1,
    ):

        for col in range(
            1,
            min(ws.max_column, 5) + 1,
        ):

            text = normalize_text(
                ws.cell(
                    row=row,
                    column=col,
                ).value
            )

            if text == wanted:
                return row

    raise RuntimeError(
        "Could not find target row:\n"
        f"{TARGET_ROW_TEXT}"
    )


# ============================================================
# FIND DATE ROW
# ============================================================

def find_date_row(
    ws,
    target_row: int,
) -> int:

    best_row = None
    best_count = 0

    for row in range(
        1,
        target_row,
    ):

        count = 0

        for col in range(
            1,
            ws.max_column + 1,
        ):

            d = parse_excel_date(
                ws.cell(
                    row=row,
                    column=col,
                ).value
            )

            if d is not None:
                count += 1

        if count > best_count:
            best_count = count
            best_row = row

    if best_row is None or best_count < 4:

        raise RuntimeError(
            "Could not detect date-header row."
        )

    return best_row


# ============================================================
# READ ONE SERIES
# ============================================================

def read_series(
    ws,
    value_name: str,
) -> pd.DataFrame:

    target_row = find_target_row(ws)

    date_row = find_date_row(
        ws,
        target_row,
    )

    print(
        f"\nSheet: {ws.title}"
    )

    print(
        "Target row:",
        target_row,
    )

    print(
        "Date row:",
        date_row,
    )

    rows = []

    for col in range(
        1,
        ws.max_column + 1,
    ):

        d = parse_excel_date(
            ws.cell(
                row=date_row,
                column=col,
            ).value
        )

        if d is None:
            continue

        value = as_number(
            ws.cell(
                row=target_row,
                column=col,
            ).value
        )

        if value is None:
            continue

        rows.append(
            {
                "date": d,
                value_name: value,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:

        raise RuntimeError(
            f"No data extracted from sheet '{ws.title}'."
        )

    return (
        df
        .sort_values("date")
        .drop_duplicates(
            "date",
            keep="last",
        )
        .reset_index(drop=True)
    )


# ============================================================
# FIND SHEETS
# ============================================================

def find_stock_sheet(wb):

    for ws in wb.worksheets:

        title = normalize_text(
            ws.title
        )

        if (
            "млн" in title
            or "руб" in title
        ):
            return ws

    return wb.worksheets[0]


def find_qoq_ex_fx_sheet(wb):

    for ws in wb.worksheets:

        title = normalize_text(
            ws.title
        )

        if (
            "qoq" in title
            and "ивп" in title
        ):
            return ws

    raise RuntimeError(
        "Sheet QOQ_ИВП was not found."
    )

def find_yoy_ex_fx_sheet(wb):

    for ws in wb.worksheets:

        title = normalize_text(
            ws.title
        )

        if (
            "yoy" in title
            and "ивп" in title
        ):
            return ws

    raise RuntimeError(
        "Sheet YOY_ИВП was not found."
    )

# ============================================================
# CALCULATE EX-FX CHANGE
# ============================================================

def calculate_ex_fx_change(
    stocks: pd.DataFrame,
    qoq: pd.DataFrame,
) -> pd.DataFrame:

    # --------------------------------------------------------
    # qoq rate dated 01.04.2026 describes growth
    # from 01.01.2026 to 01.04.2026.
    #
    # Therefore:
    #
    # change_t =
    # previous actual debt stock
    # * qoq_ex_fx_t / 100
    # --------------------------------------------------------

    stocks = stocks.copy()

    stocks[
        "previous_actual_debt_mln_rub"
    ] = (
        stocks[
            "total_debt_mln_rub"
        ]
        .shift(1)
    )

    result = stocks.merge(
        qoq,
        on="date",
        how="left",
    )

    result[
        "debt_change_ex_fx_mln_rub"
    ] = (
        result[
            "previous_actual_debt_mln_rub"
        ]
        * result[
            "qoq_growth_ex_fx_pct"
        ]
        / 100
    )

    return result


# ============================================================
# VALIDATION
# ============================================================

def validate(
    df: pd.DataFrame,
) -> None:

    if df.empty:

        raise RuntimeError(
            "Debt output is empty."
        )

    if df[
        "date"
    ].duplicated().any():

        raise RuntimeError(
            "Duplicate dates detected."
        )

    if df[
        "total_debt_mln_rub"
    ].isna().any():

        raise RuntimeError(
            "Missing total debt values."
        )

    missing_rates = df[
        df["date"] > df["date"].min()
    ][
        "qoq_growth_ex_fx_pct"
    ].isna()

    if missing_rates.any():

        bad_rows = df[
            df["date"] > df["date"].min()
        ][
            missing_rates
        ]

        raise RuntimeError(
            "Missing QOQ_ИВП values:\n"
            + bad_rows.to_string(
                index=False
            )
        )

    dates = (
        df[
            "date"
        ]
        .sort_values()
        .reset_index(
            drop=True
        )
    )

    for i in range(
        1,
        len(dates),
    ):

        previous = dates.iloc[
            i - 1
        ]

        current = dates.iloc[
            i
        ]

        months = (
            (current.year - previous.year) * 12
            + current.month
            - previous.month
        )

        if months != 3:

            raise RuntimeError(
                "Debt series is not continuous quarterly.\n"
                f"Previous: {previous.date()}\n"
                f"Current: {current.date()}"
            )

    print(
        "\nLatest available date:",
        dates.max().strftime(
            "%Y-%m-%d"
        ),
    )
    missing_yoy_rates = df[
        df["date"] >= pd.Timestamp("2021-01-01")
        ][
        "yoy_growth_ex_fx_pct"
    ].isna()

    if missing_yoy_rates.any():
        bad_rows = df[
            df["date"] >= pd.Timestamp("2021-01-01")
            ][
            missing_yoy_rates
        ]

        raise RuntimeError(
            "Missing YOY_ИВП values:\n"
            + bad_rows.to_string(
                index=False
            )
        )


# ============================================================
# MAIN
# ============================================================

def main():

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    content = download_workbook()

    wb = load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=False,
    )

    if not wb.worksheets:

        raise RuntimeError(
            "Workbook contains no worksheets."
        )

    # --------------------------------------------------------
    # 1. Actual debt stocks
    # --------------------------------------------------------

    stock_ws = find_stock_sheet(
        wb
    )

    stocks = read_series(
        stock_ws,
        "total_debt_mln_rub",
    )

    # --------------------------------------------------------
    # 2. QoQ growth excluding FX revaluation
    # --------------------------------------------------------

    qoq_ws = find_qoq_ex_fx_sheet(
        wb
    )

    qoq = read_series(
        qoq_ws,
        "qoq_growth_ex_fx_pct",
    )

    # --------------------------------------------------------
    # 3. YoY growth excluding FX revaluation
    # --------------------------------------------------------

    yoy_ws = find_yoy_ex_fx_sheet(
        wb
    )

    yoy = read_series(
        yoy_ws,
        "yoy_growth_ex_fx_pct",
    )

    print(
        "\nRaw YOY_ИВП range:"
    )

    print(
        yoy["date"].min().date(),
        "->",
        yoy["date"].max().date(),
    )
    print(
        "\nRaw stock range:"
    )

    print(
        stocks["date"].min().date(),
        "->",
        stocks["date"].max().date(),
    )

    print(
        "\nRaw QOQ_ИВП range:"
    )

    print(
        qoq["date"].min().date(),
        "->",
        qoq["date"].max().date(),
    )

    # --------------------------------------------------------
    # 3. Calculate quarterly change excluding FX revaluation
    # --------------------------------------------------------

    debt = calculate_ex_fx_change(
        stocks,
        qoq,
    )

    debt = debt.merge(
        yoy,
        on="date",
        how="left",
    )

    # --------------------------------------------------------
    # 4. Keep output from 2020 onward
    # --------------------------------------------------------

    debt = debt[
        debt["date"]
        >= FIRST_OUTPUT_DATE
    ].copy()

    validate(
        debt
    )

    # --------------------------------------------------------
    # 5. Add useful comparison column
    # --------------------------------------------------------

    debt[
        "debt_change_ex_fx_trln_rub"
    ] = (
        debt[
            "debt_change_ex_fx_mln_rub"
        ]
        / 1_000_000
    )

    # --------------------------------------------------------
    # 6. Source
    # --------------------------------------------------------

    debt[
        "source"
    ] = SOURCE_NAME

    # --------------------------------------------------------
    # 7. Diagnostics against CBR page
    # --------------------------------------------------------

    comparison_dates = [
        pd.Timestamp("2026-01-01"),
        pd.Timestamp("2026-04-01"),
    ]

    comparison = debt[
        debt["date"].isin(
            comparison_dates
        )
    ][
        [
            "date",
            "previous_actual_debt_mln_rub",
            "qoq_growth_ex_fx_pct",
            "debt_change_ex_fx_mln_rub",
            "debt_change_ex_fx_trln_rub",
            "total_debt_mln_rub",
        ]
    ].copy()

    print(
        "\n" + "=" * 72
    )

    print(
        "CBR CHECK"
    )

    print(
        "=" * 72
    )

    print(
        "\nExpected approximately:"
    )

    print(
        "2026-01-01 -> 5.5 trln RUB"
    )

    print(
        "2026-04-01 -> 1.1 trln RUB"
    )

    print(
        "\nCalculated:"
    )

    if not comparison.empty:

        comparison_display = (
            comparison.copy()
        )

        comparison_display[
            "date"
        ] = (
            comparison_display[
                "date"
            ]
            .dt.strftime(
                "%Y-%m-%d"
            )
        )

        print(
            comparison_display
            .to_string(
                index=False
            )
        )

    # --------------------------------------------------------
    # 8. Final CSV
    # --------------------------------------------------------

    debt[
        "total_debt_mln_rub"
    ] = (
        debt[
            "total_debt_mln_rub"
        ]
        .round(0)
        .astype("int64")
    )

    debt[
        "qoq_growth_ex_fx_pct"
    ] = (
        debt[
            "qoq_growth_ex_fx_pct"
        ]
        .round(6)
    )
    debt[
        "yoy_growth_ex_fx_pct"
    ] = (
        debt[
            "yoy_growth_ex_fx_pct"
        ]
        .round(6)
    )

    debt[
        "debt_change_ex_fx_mln_rub"
    ] = (
        debt[
            "debt_change_ex_fx_mln_rub"
        ]
        .round(0)
        .astype("int64")
    )

    debt[
        "debt_change_ex_fx_trln_rub"
    ] = (
        debt[
            "debt_change_ex_fx_trln_rub"
        ]
        .round(6)
    )

    debt[
        "date"
    ] = (
        debt[
            "date"
        ]
        .dt.strftime(
            "%Y-%m-%d"
        )
    )

    debt = debt[
        [
            "date",
            "total_debt_mln_rub",
            "qoq_growth_ex_fx_pct",
            "yoy_growth_ex_fx_pct",
            "debt_change_ex_fx_mln_rub",
            "debt_change_ex_fx_trln_rub",
            "source",
        ]
    ]

    debt.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    # --------------------------------------------------------
    # 9. Source info
    # --------------------------------------------------------

    SOURCE_INFO_FILE.write_text(
        "\n".join(
            [
                "Russia extended debt",
                "source=Bank of Russia",
                f"source_page={SOURCE_PAGE}",
                f"source_file={SOURCE_FILE_URL}",
                (
                    "series="
                    "Долг организаций и домашних хозяйств, итого"
                ),
                "unit_stock=mln_rub",
                "unit_change=mln_rub",
                (
                    "qoq_series="
                    "QOQ_ИВП, quarterly growth excluding "
                    "foreign-exchange revaluation"
                ),
                (
                    "yoy_series="
                    "YOY_ИВП, annual growth excluding "
                    "foreign-exchange revaluation"
                ),
                (
                    "change_formula="
                    "previous actual debt stock "
                    "* qoq_growth_ex_fx_pct / 100"
                ),
                (
                    "change_definition="
                    "quarterly debt change excluding "
                    "foreign-exchange revaluation"
                ),
                "output_start=2020-01-01",
                f"output_end={debt['date'].max()}",
                "",
            ]
        ),
        encoding="utf-8",
    )

    # --------------------------------------------------------
    # 10. Final diagnostics
    # --------------------------------------------------------

    print(
        "\n" + "=" * 72
    )

    print(
        "DONE"
    )

    print(
        "=" * 72
    )

    print(
        "\nOutput:"
    )

    print(
        OUTPUT_FILE
    )

    print(
        "\nSource info:"
    )

    print(
        SOURCE_INFO_FILE
    )

    print(
        "\nRange:"
    )

    print(
        debt["date"].min(),
        "->",
        debt["date"].max(),
    )

    print(
        "\nLast rows:"
    )

    print(
        debt
        .tail(12)
        .to_string(
            index=False
        )
    )


if __name__ == "__main__":
    main()